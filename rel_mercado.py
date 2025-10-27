import sys
from forms.tela_rel_mercado import *
from banco_dados.conexao_nuvem import conectar_banco_nuvem
from banco_dados.controle_erros import grava_erro_banco
from comandos.telas import tamanho_aplicacao, icone
from comandos.tabelas import layout_cabec_tab, lanca_tabela, extrair_tabela
from comandos.conversores import valores_para_float, valores_para_virgula, float_para_moeda_reais
from PyQt5.QtWidgets import QMainWindow, QApplication, QMessageBox
from PyQt5.QtChart import QChart, QChartView, QPieSeries, QPieSlice
from PyQt5.QtGui import QPainter, QColor, QFont
from PyQt5.QtCore import Qt
import inspect
import os
from datetime import datetime
import traceback


class TelaRelatorioMercado(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        self.processando = False

        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "gremio.png")
        tamanho_aplicacao(self)

        layout_cabec_tab(self.table_Lista)

        self.id_usuario = "1"

        self.lanca_combo_grupo()
        self.configurar_data_atual()

        self.combo_Grupo.activated.connect(self.lanca_combo_produto)
        self.btn_Consulta.clicked.connect(self.manipula_dados)
        self.btn_Excel.clicked.connect(self.gerar_excel)

        self.line_Ano.editingFinished.connect(self.manipula_dados)

    def mensagem_alerta(self, mensagem):
        try:
            alert = QMessageBox()
            alert.setIcon(QMessageBox.Warning)
            alert.setText(mensagem)
            alert.setWindowTitle("Atenção")
            alert.setStandardButtons(QMessageBox.Ok)
            alert.exec_()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            self.mensagem_alerta(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                                 f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                                 f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def configurar_data_atual(self):
        try:
            # Obter o mês e o ano atuais
            data_atual = datetime.now()
            mes_atual = data_atual.month  # Exemplo: 11 para novembro
            ano_atual = data_atual.year  # Exemplo: 2024

            # Encontrar o índice do mês atual no combo box e selecioná-lo
            for i in range(self.combo_Meses.count()):
                item_text = self.combo_Meses.itemText(i)
                if item_text.startswith(f"{mes_atual} -"):
                    self.combo_Meses.setCurrentIndex(i)
                    break

            # Configurar o ano atual no QLineEdit
            self.line_Ano.setText(str(ano_atual))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_combo_grupo(self):
        conecta = conectar_banco_nuvem()
        try:
            self.combo_Grupo.clear()

            nova_lista = [""]

            cursor = conecta.cursor()
            cursor.execute('SELECT id, descricao FROM cadastro_gr_produto_mercado order by descricao;')
            lista_completa = cursor.fetchall()
            for ides, descr in lista_completa:
                dd = f"{ides} - {descr}"
                nova_lista.append(dd)

            self.combo_Grupo.addItems(nova_lista)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def lanca_combo_produto(self):
        conecta = conectar_banco_nuvem()
        try:
            self.combo_Produtos.clear()

            nova_lista = [""]

            grupo = self.combo_Grupo.currentText()
            if grupo:
                grupotete = grupo.find(" - ")
                id_grupo = grupo[:grupotete]

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, descricao "
                               f"FROM cadastro_produto_mercado "
                               f"where id_gr_produto = '{id_grupo}' "
                               f"order by descricao;")
                lista_completa = cursor.fetchall()
                for ides, descr in lista_completa:
                    dd = f"{ides} - {descr}"
                    nova_lista.append(dd)

                self.combo_Produtos.addItems(nova_lista)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def manipula_dados(self):
        conecta = conectar_banco_nuvem()

        if not self.processando:
            try:
                self.processando = True

                meses = self.combo_Meses.currentText()
                ano = self.line_Ano.text()

                grupo = self.combo_Grupo.currentText()

                produto = self.combo_Produtos.currentText()

                classifica = self.combo_Classifica.currentText()

                if classifica == "MAIOR VALOR":
                    order_by = "ORDER BY TOTAL DESC"
                elif classifica == "MENOR VALOR":
                    order_by = "ORDER BY TOTAL"
                elif classifica == "MAIOR UNIT":
                    order_by = "ORDER BY comp.unit DESC"
                elif classifica == "MENOR UNIT":
                    order_by = "ORDER BY comp.unit"
                elif classifica == "GRUPO":
                    order_by = "ORDER BY grup.descricao"
                elif classifica == "PRODUTO":
                    order_by = "ORDER BY prod.descricao"
                elif classifica == "ESTABELECIMENTO":
                    order_by = "ORDER BY estab.descricao"
                else:
                    order_by = "ORDER BY mov.data"

                if meses and ano:
                    if not grupo:
                        meses_tete = meses.find(" - ")
                        num_mes = meses[:meses_tete]

                        ano_int = int(ano)

                        self.total_mensal(num_mes, ano_int, order_by)
                        self.adicionar_grafico(num_mes, ano_int)
                    else:
                        meses_tete = meses.find(" - ")
                        num_mes = meses[:meses_tete]

                        ano_int = int(ano)

                        grupo_tete = grupo.find(" - ")
                        num_grupo = grupo[:grupo_tete]

                        produto_tete = produto.find(" - ")
                        num_produto = produto[:produto_tete]

                        if produto:
                            self.total_mensal_produto(num_mes, ano_int, num_produto, order_by)
                            self.adicionar_grafico(num_mes, ano_int)
                        else:
                            self.total_mensal_grupo(num_mes, ano_int, num_grupo, order_by)
                            self.adicionar_grafico(num_mes, ano_int)

                elif ano and not meses:
                    if not grupo:
                        ano_int = int(ano)

                        self.total_anual(ano_int, order_by)
                        self.adicionar_grafico_ano(ano_int)
                    else:
                        ano_int = int(ano)

                        grupo_tete = grupo.find(" - ")
                        num_grupo = grupo[:grupo_tete]

                        produto_tete = produto.find(" - ")
                        num_produto = produto[:produto_tete]

                        if produto:
                            self.total_anual_produto(ano_int, num_produto, order_by)
                            self.adicionar_grafico_ano(ano_int)
                        else:
                            print("entrei no grupo")
                            self.total_anual_grupo(ano_int, num_grupo, order_by)
                            self.adicionar_grafico_ano(ano_int)

            except Exception as e:
                nome_funcao = inspect.currentframe().f_code.co_name
                exc_traceback = sys.exc_info()[2]
                self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

            finally:
                self.processando = False
                if 'conexao' in locals():
                    conecta.close()

    def select_padrao(self):
        try:
            texto_padrao = """SELECT DATE_FORMAT(mov.data, '%d/%m/%Y'), estab.descricao, 
                            grup.descricao, prod.descricao, prod.um, comp.qtde, comp.unit, 
                            (comp.qtde * comp.unit) AS TOTAL, comp.obs 
                           FROM compras_mercado as comp 
                           INNER JOIN cadastro_produto_mercado AS prod ON comp.id_produto = prod.id 
                           INNER JOIN movimentacao AS mov ON comp.id_movimentacao = mov.id 
                           INNER JOIN cadastro_gr_produto_mercado as grup ON prod.id_gr_produto = grup.id 
                           INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id """

            return texto_padrao

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def total_mensal(self, num_mes, ano_int, order_by):
        conecta = conectar_banco_nuvem()
        try:
            self.table_Lista.setRowCount(0)

            nova_lista_tabela = []

            texto_padrao = self.select_padrao()

            cursor = conecta.cursor()
            cursor.execute(f"""{texto_padrao}
                        WHERE MONTH(mov.data) = {num_mes}
                          AND YEAR(mov.data) = {ano_int} 
                          {order_by};
                    """)
            lista_completa = cursor.fetchall()

            total_final = 0
            qtde_total = 0

            if lista_completa:
                for i in lista_completa:
                    data, estab, grupo, produto, um, qtde, unit, total, obs = i

                    qtde_float = valores_para_float(qtde)
                    unit_float = valores_para_float(unit)

                    qtde_virg = valores_para_virgula(qtde)

                    unit_moeda = float_para_moeda_reais(unit_float)
                    total_moeda = float_para_moeda_reais(total)

                    total_final += total
                    qtde_total += qtde_float

                    dados = (data, grupo, produto, um, qtde_virg, unit_moeda, total_moeda, estab, obs)
                    nova_lista_tabela.append(dados)

            if nova_lista_tabela:
                lanca_tabela(self.table_Lista, nova_lista_tabela)

            if total_final:
                total_arred = float_para_moeda_reais(round(total_final, 2))
                qtde_arred = valores_para_virgula(round(qtde_total, 2))

                self.label_RS_Total.setText(str(total_arred))
                self.label_Qtde_Total.setText(str(qtde_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def total_mensal_grupo(self, num_mes, ano_int, num_grupo, order_by):
        conecta = conectar_banco_nuvem()
        try:
            self.table_Lista.setRowCount(0)

            nova_lista_tabela = []

            texto_padrao = self.select_padrao()

            cursor = conecta.cursor()
            cursor.execute(f"""{texto_padrao}
                        WHERE MONTH(mov.data) = {num_mes}
                          AND YEAR(mov.data) = {ano_int} 
                          AND grup.id  = {num_grupo} 
                        {order_by};
                    """)
            lista_completa = cursor.fetchall()

            total_final = 0
            qtde_total = 0

            if lista_completa:
                for i in lista_completa:
                    data, estab, grupo, produto, um, qtde, unit, total, obs = i

                    qtde_float = valores_para_float(qtde)
                    unit_float = valores_para_float(unit)

                    qtde_virg = valores_para_virgula(qtde)

                    unit_moeda = float_para_moeda_reais(unit_float)
                    total_moeda = float_para_moeda_reais(total)

                    total_final += total
                    qtde_total += qtde_float

                    dados = (data, grupo, produto, um, qtde_virg, unit_moeda, total_moeda, estab, obs)
                    nova_lista_tabela.append(dados)

            if nova_lista_tabela:
                lanca_tabela(self.table_Lista, nova_lista_tabela)

            if total_final:
                total_arred = float_para_moeda_reais(round(total_final, 2))
                qtde_arred = valores_para_virgula(round(qtde_total, 2))

                self.label_RS_Total.setText(str(total_arred))
                self.label_Qtde_Total.setText(str(qtde_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def total_mensal_produto(self, num_mes, ano_int, num_produto, order_by):
        conecta = conectar_banco_nuvem()
        try:
            self.table_Lista.setRowCount(0)

            nova_lista_tabela = []

            texto_padrao = self.select_padrao()

            cursor = conecta.cursor()
            cursor.execute(f"""{texto_padrao}
                        WHERE MONTH(mov.data) = {num_mes}
                          AND YEAR(mov.data) = {ano_int} 
                          AND prod.id = {num_produto} 
                        {order_by};
                    """)
            lista_completa = cursor.fetchall()

            total_final = 0
            qtde_total = 0

            if lista_completa:
                for i in lista_completa:
                    data, estab, grupo, produto, um, qtde, unit, total, obs = i

                    qtde_float = valores_para_float(qtde)
                    unit_float = valores_para_float(unit)

                    qtde_virg = valores_para_virgula(qtde)

                    unit_moeda = float_para_moeda_reais(unit_float)
                    total_moeda = float_para_moeda_reais(total)

                    total_final += total
                    qtde_total += qtde_float

                    dados = (data, grupo, produto, um, qtde_virg, unit_moeda, total_moeda, estab, obs)
                    nova_lista_tabela.append(dados)

            if nova_lista_tabela:
                lanca_tabela(self.table_Lista, nova_lista_tabela)

            if total_final:
                total_arred = float_para_moeda_reais(round(total_final, 2))
                qtde_arred = valores_para_virgula(round(qtde_total, 2))

                self.label_RS_Total.setText(str(total_arred))
                self.label_Qtde_Total.setText(str(qtde_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def adicionar_grafico(self, num_mes, ano_int):
        conecta = conectar_banco_nuvem()
        try:
            # Criar a série de dados
            series = QPieSeries()

            cursor = conecta.cursor()
            cursor.execute(f"""
                SELECT grup.descricao, 
                       SUM(comp.qtde * comp.unit) AS total_valor
                FROM compras_mercado as comp 
                INNER JOIN cadastro_produto_mercado AS prod ON comp.id_produto = prod.id 
               INNER JOIN movimentacao AS mov ON comp.id_movimentacao = mov.id 
               INNER JOIN cadastro_gr_produto_mercado as grup ON prod.id_gr_produto = grup.id 
               INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id
                WHERE MONTH(mov.data) = {num_mes}
                  AND YEAR(mov.data) = {ano_int}  
                GROUP BY grup.descricao
                ORDER BY total_valor DESC;
            """)
            lista_totais = cursor.fetchall()
            if lista_totais:
                # Definir uma paleta de cores
                cores = [
                    QColor("#FF5733"), QColor("#33FF57"), QColor("#3357FF"),
                    QColor("#FF33A1"), QColor("#A133FF"), QColor("#FF8333"),
                    QColor("#33FFF5"), QColor("#F533FF")
                ]

                total = 0

                for i, (categoria, valor) in enumerate(lista_totais):
                    valor_float = valores_para_float(valor)

                    total += valor_float

                    fatia = series.append(categoria, valor_float)

                    # Atribuir uma cor da paleta, usando o índice de forma cíclica
                    fatia.setBrush(cores[i % len(cores)])

                    # Exibir rótulos fora da fatia com descrição e valor
                    fatia.setLabelPosition(QPieSlice.LabelOutside)
                    fatia.setLabel(f"{categoria}: {valor_float:.2f}")
                    fatia.setLabelVisible(True)
                    fatia.setLabelFont(QFont("Arial", 6))  # Define fonte menor (8pt)

                    if valor_float / total < 0.05:  # menos de 5%
                        fatia.setLabelVisible(False)

                # <<< AQUI, depois de adicionar todas as fatias, ajusta o diâmetro
                series.setPieSize(0.4)  # 70% do espaço do widget

                # Criar o gráfico
                chart = QChart()
                chart.addSeries(series)

                # Configurar a legenda
                chart.legend().setVisible(True)  # Tornar a legenda visível
                chart.legend().setAlignment(Qt.AlignRight)  # Posicionar a legenda à direita
                chart.legend().setFont(QFont("Arial", 7))

                # Configurar o ChartView
                chart_view = QChartView(chart)
                chart_view.setRenderHint(QPainter.Antialiasing)

                # Adicionar o ChartView ao widget da interface
                layout = self.widget_grafico.layout()  # Obtém o layout do widget
                if not layout:  # Se o layout não existir, cria um
                    from PyQt5.QtWidgets import QVBoxLayout
                    layout = QVBoxLayout(self.widget_grafico)
                else:  # Remove gráficos antigos do layout
                    for i in reversed(range(layout.count())):
                        layout.itemAt(i).widget().deleteLater()

                layout.addWidget(chart_view)  # Adiciona o gráfico ao layout

                if total:
                    total_arred = float_para_moeda_reais(round(total, 2))
                    self.label_Total_Grafico.setText(str(total_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def total_anual(self, ano_int, order_by):
        conecta = conectar_banco_nuvem()
        try:
            self.table_Lista.setRowCount(0)

            nova_lista_tabela = []

            texto_padrao = self.select_padrao()

            cursor = conecta.cursor()
            cursor.execute(f"""{texto_padrao}
                        WHERE YEAR(mov.data) = {ano_int} 
                        {order_by};
                    """)
            lista_completa = cursor.fetchall()

            total_final = 0
            qtde_total = 0

            if lista_completa:
                for i in lista_completa:
                    data, estab, grupo, produto, um, qtde, unit, total, obs = i

                    qtde_float = valores_para_float(qtde)
                    unit_float = valores_para_float(unit)

                    qtde_virg = valores_para_virgula(qtde)

                    unit_moeda = float_para_moeda_reais(unit_float)
                    total_moeda = float_para_moeda_reais(total)

                    total_final += total
                    qtde_total += qtde_float

                    dados = (data, grupo, produto, um, qtde_virg, unit_moeda, total_moeda, estab, obs)
                    nova_lista_tabela.append(dados)

            if nova_lista_tabela:
                lanca_tabela(self.table_Lista, nova_lista_tabela)

            if total_final:
                total_arred = float_para_moeda_reais(round(total_final, 2))
                qtde_arred = valores_para_virgula(round(qtde_total, 2))

                self.label_RS_Total.setText(str(total_arred))
                self.label_Qtde_Total.setText(str(qtde_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def total_anual_grupo(self, ano_int, num_grupo, order_by):
        conecta = conectar_banco_nuvem()
        try:
            self.table_Lista.setRowCount(0)

            nova_lista_tabela = []

            texto_padrao = self.select_padrao()

            cursor = conecta.cursor()
            cursor.execute(f"""{texto_padrao}
                        WHERE YEAR(mov.data) = {ano_int} 
                          AND grup.id = {num_grupo} 
                        {order_by};
                    """)
            lista_completa = cursor.fetchall()

            total_final = 0
            qtde_total = 0

            if lista_completa:
                for i in lista_completa:
                    data, estab, grupo, produto, um, qtde, unit, total, obs = i

                    qtde_float = valores_para_float(qtde)
                    unit_float = valores_para_float(unit)

                    qtde_virg = valores_para_virgula(qtde)

                    unit_moeda = float_para_moeda_reais(unit_float)
                    total_moeda = float_para_moeda_reais(total)

                    total_final += total
                    qtde_total += qtde_float

                    dados = (data, grupo, produto, um, qtde_virg, unit_moeda, total_moeda, estab, obs)
                    nova_lista_tabela.append(dados)

            if nova_lista_tabela:
                lanca_tabela(self.table_Lista, nova_lista_tabela)

            if total_final:
                total_arred = float_para_moeda_reais(round(total_final, 2))
                qtde_arred = valores_para_virgula(round(qtde_total, 2))

                self.label_RS_Total.setText(str(total_arred))
                self.label_Qtde_Total.setText(str(qtde_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def total_anual_produto(self, ano_int, num_produto, order_by):
        conecta = conectar_banco_nuvem()
        try:
            self.table_Lista.setRowCount(0)

            nova_lista_tabela = []

            texto_padrao = self.select_padrao()

            cursor = conecta.cursor()
            cursor.execute(f"""{texto_padrao}
                        WHERE YEAR(mov.data) = {ano_int} 
                          AND prod.id = {num_produto}  
                        {order_by};
                    """)
            lista_completa = cursor.fetchall()

            total_final = 0
            qtde_total = 0

            if lista_completa:
                for i in lista_completa:
                    data, estab, grupo, produto, um, qtde, unit, total, obs = i

                    qtde_float = valores_para_float(qtde)
                    unit_float = valores_para_float(unit)

                    qtde_virg = valores_para_virgula(qtde)

                    unit_moeda = float_para_moeda_reais(unit_float)
                    total_moeda = float_para_moeda_reais(total)

                    total_final += total
                    qtde_total += qtde_float

                    dados = (data, grupo, produto, um, qtde_virg, unit_moeda, total_moeda, estab, obs)
                    nova_lista_tabela.append(dados)

            if nova_lista_tabela:
                lanca_tabela(self.table_Lista, nova_lista_tabela)

            if total_final:
                total_arred = float_para_moeda_reais(round(total_final, 2))
                qtde_arred = valores_para_virgula(round(qtde_total, 2))

                self.label_RS_Total.setText(str(total_arred))
                self.label_Qtde_Total.setText(str(qtde_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def adicionar_grafico_ano(self, ano_int):
        conecta = conectar_banco_nuvem()
        try:
            # Criar a série de dados
            series = QPieSeries()

            cursor = conecta.cursor()
            cursor.execute(f"""
                SELECT grup.descricao, 
                       SUM(comp.qtde * comp.unit) AS total_valor
                FROM compras_mercado as comp 
                INNER JOIN cadastro_produto_mercado AS prod ON comp.id_produto = prod.id 
               INNER JOIN movimentacao AS mov ON comp.id_movimentacao = mov.id 
               INNER JOIN cadastro_gr_produto_mercado as grup ON prod.id_gr_produto = grup.id 
               INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id
                WHERE YEAR(mov.data) = {ano_int}
                GROUP BY grup.descricao
                ORDER BY total_valor DESC;
            """)
            lista_totais = cursor.fetchall()
            if lista_totais:
                # Definir uma paleta de cores
                cores = [
                    QColor("#FF5733"), QColor("#33FF57"), QColor("#3357FF"),
                    QColor("#FF33A1"), QColor("#A133FF"), QColor("#FF8333"),
                    QColor("#33FFF5"), QColor("#F533FF")
                ]

                total = 0

                for i, (categoria, valor) in enumerate(lista_totais):
                    valor_float = valores_para_float(valor)

                    total += valor_float

                    fatia = series.append(categoria, valor_float)

                    # Atribuir uma cor da paleta, usando o índice de forma cíclica
                    fatia.setBrush(cores[i % len(cores)])

                    # Exibir rótulos fora da fatia com descrição e valor
                    fatia.setLabelPosition(QPieSlice.LabelOutside)
                    fatia.setLabel(f"{categoria}: {valor_float:.2f}")
                    fatia.setLabelVisible(True)
                    fatia.setLabelFont(QFont("Arial", 6))  # Define fonte menor (8pt)

                    if valor_float / total < 0.05:  # menos de 5%
                        fatia.setLabelVisible(False)

                # <<< AQUI, depois de adicionar todas as fatias, ajusta o diâmetro
                series.setPieSize(0.4)  # 70% do espaço do widget

                # Criar o gráfico
                chart = QChart()
                chart.addSeries(series)

                # Configurar a legenda
                chart.legend().setVisible(True)  # Tornar a legenda visível
                chart.legend().setAlignment(Qt.AlignRight)  # Posicionar a legenda à direita
                chart.legend().setFont(QFont("Arial", 7))

                # Configurar o ChartView
                chart_view = QChartView(chart)
                chart_view.setRenderHint(QPainter.Antialiasing)

                # Adicionar o ChartView ao widget da interface
                layout = self.widget_grafico.layout()  # Obtém o layout do widget
                if not layout:  # Se o layout não existir, cria um
                    from PyQt5.QtWidgets import QVBoxLayout
                    layout = QVBoxLayout(self.widget_grafico)
                else:  # Remove gráficos antigos do layout
                    for i in reversed(range(layout.count())):
                        layout.itemAt(i).widget().deleteLater()

                layout.addWidget(chart_view)  # Adiciona o gráfico ao layout

                if total:
                    total_arred = round(total, 2)
                    self.label_Total_Grafico.setText(str(total_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def gerar_excel(self):
        conecta = conectar_banco_nuvem()
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            dados_tabela = extrair_tabela(self.table_Lista)

            if dados_tabela:
                meses = self.combo_Meses.currentText()
                ano = self.line_Ano.text()

                meses_tete = meses.find(" - ")
                num_mes = meses[:meses_tete]

                # Caminho para salvar na área de trabalho
                desktop = os.path.join(os.path.expanduser("~"), "Desktop")
                arquivo_excel = os.path.join(desktop, f"Despesa Mercado {num_mes} de {ano}.xlsx")

                # Cria um novo workbook e planilha
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Mercado"

                # Estilos
                bold_font = Font(bold=True, color="FFFFFF")
                fill_cinza = PatternFill("solid", fgColor="808080")
                alinhamento_centro = Alignment(horizontal="center", vertical="center")
                borda_preta = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000")
                )

                # Cabeçalho
                headers = ["Data", "Grupo", "Produto", "UM", "Qtde", "R$/Unid", "R$", "Estabelecimento", "Observação"]
                ws.append(headers)

                # Aplicar estilo no cabeçalho
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = bold_font
                    cell.fill = fill_cinza
                    cell.alignment = alinhamento_centro
                    cell.border = borda_preta

                # Adiciona os dados
                for linha_idx, (data, grupo, produto, um, qtde, unit, total, estab, obs) in (
                        enumerate(dados_tabela, start=2)):
                    qtde_float = valores_para_float(qtde)
                    unit_float = valores_para_float(unit)
                    total_float = valores_para_float(total)

                    ws.append([data, grupo, produto, um, qtde_float, unit_float, total_float, estab, obs])

                    # Formata quantidade com 3 casas decimais (coluna 5)
                    qtde_cell = ws.cell(row=linha_idx, column=5)
                    qtde_cell.number_format = '0.000'

                    # Formata valores monetários (colunas 6 e 7)
                    unit_cell = ws.cell(row=linha_idx, column=6)
                    total_cell = ws.cell(row=linha_idx, column=7)
                    unit_cell.number_format = 'R$ #,##0.00'  # formato brasileiro
                    total_cell.number_format = 'R$ #,##0.00'

                # Aplica borda e alinhamento a todos os dados
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                    for cell in row:
                        cell.border = borda_preta
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                # Ajustar largura das colunas automaticamente
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_length + 2

                # Salva o arquivo na área de trabalho
                wb.save(arquivo_excel)
                self.label_Excel.setText(f"Excel Salvo!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conecta' in locals():
                conecta.close()


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaRelatorioMercado()
    tela.show()
    qt.exec_()
