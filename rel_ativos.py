import sys
from forms.tela_rel_ativos import *
from banco_dados.conexao_nuvem import conectar_banco_nuvem
from banco_dados.controle_erros import grava_erro_banco
from comandos.telas import tamanho_aplicacao, icone
from comandos.tabelas import layout_cabec_tab, lanca_tabela, extrair_tabela
from comandos.conversores import float_para_moeda_reais, valores_para_float, data_banco_para_brasileiro
from PyQt5.QtWidgets import QMainWindow, QApplication, QMessageBox
import inspect
import os
import traceback

import yfinance as yf
from decimal import Decimal


class TelaRelatorioAtivos(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        self.processando = False

        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "gremio.png")
        tamanho_aplicacao(self)

        layout_cabec_tab(self.table_Lista)

        self.btn_Excel.clicked.connect(self.gerar_excel)

        nova_lista = self.lanca_dados_fixo()
        self.lanca_dados_variavel(nova_lista)
        self.calcular_totais()

    def mensagem_alerta(self, mensagem):
        try:
            alert = QMessageBox()
            alert.setIcon(QMessageBox.Warning)
            alert.setText(mensagem)
            alert.setWindowTitle("Atenção")
            alert.setStandardButtons(QMessageBox.Ok)
            alert.exec_()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            self.mensagem_alerta(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                                 f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                                 f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def lanca_dados_fixo(self):
        conecta = conectar_banco_nuvem()
        try:
            nova_lista = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT ativ.id, mov.data, banc.DESCRICAO, ativ.NOME_TITULO, indexa.DESCRICAO, "
                           f"tip.DESCRICAO, ativ.RENTABILIDADE, "
                           f"ativ.vencimento, ativ.LIQUIDEZ, ativ.VALOR, ativ.status "
                           f"FROM fixa_cadastro_ativo as ativ "
                           f"INNER JOIN movimentacao as mov ON ativ.id_movimentacao = mov.id "
                           f"INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id "
                           f"INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id "
                           f"INNER JOIN fixa_cadastro_indexador AS indexa ON ativ.ID_INDEXADOR = indexa.id "
                           f"INNER JOIN fixa_cadastro_tipo AS tip ON ativ.ID_TIPO_FIXA = tip.id "
                           f"where ativ.status = 'A' order by mov.data;")
            dados_ativo = cursor.fetchall()

            if dados_ativo:
                for i in dados_ativo:
                    id_ativo, data_mov, banco, nome, indexa, tipo, rent, venc, liq, valor, status = i

                    data_mov = data_banco_para_brasileiro(data_mov)
                    venc = data_banco_para_brasileiro(venc)

                    valor = float_para_moeda_reais(valor)

                    saldo = self.definir_saldo_fixo(id_ativo)

                    if banco == "FGTS":
                        dados = (data_mov, banco, "FGTS", nome, valor, venc, saldo, liq, status)
                    else:
                        dados = (data_mov, banco, "FIXO", nome, valor, venc, saldo, liq, status)
                    nova_lista.append(dados)

            if nova_lista:
                lanca_tabela(self.table_Lista, nova_lista)

            return nova_lista

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def definir_saldo_fixo(self, num_ativo):
        conecta = conectar_banco_nuvem()
        try:
            valor = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT mov.data, ativ.NOME_TITULO, indexa.DESCRICAO, tip.DESCRICAO, "
                           f"ativ.RENTABILIDADE, ativ.LIQUIDEZ, ativ.VALOR, ativ.status "
                           f"FROM fixa_cadastro_ativo as ativ "
                           f"INNER JOIN movimentacao as mov ON ativ.id_movimentacao = mov.id "
                           f"INNER JOIN fixa_cadastro_indexador AS indexa ON ativ.ID_INDEXADOR = indexa.id "
                           f"INNER JOIN fixa_cadastro_tipo AS tip ON ativ.ID_TIPO_FIXA = tip.id "
                           f"where ativ.id = {num_ativo};")
            dados_ativo = cursor.fetchall()
            data, nome, indexa, tipo, rent, liq, valor_compra, status = dados_ativo[0]

            valor_compra = valores_para_float(valor_compra)

            valor += valor_compra

            cursor = conecta.cursor()
            cursor.execute(f"SELECT DATE_FORMAT(mov.data, '%d/%m/%Y'), banc.descricao, mov.id, "
                           f"cat.descricao, (mov.qtde_ent - mov.qtde_sai) as valor, mov.obs "
                           f"FROM fixa_operacao as ope "
                           f"INNER JOIN movimentacao as mov ON ope.id_movimentacao = mov.id "
                           f"INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id "
                           f"INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id "
                           f"INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id "
                           f"where ope.id_ativo = {num_ativo} "
                           f"ORDER BY mov.data, valor DESC;")
            lista_completa = cursor.fetchall()

            if lista_completa:
                for i in lista_completa:
                    data, banco, num_mov, tipo, valor_operacoes, obs = i

                    valor_operacoes = valores_para_float(valor_operacoes)
                    valor += valor_operacoes

            valor = round(valor, 2)
            valor = float_para_moeda_reais(valor)

            return valor

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def lanca_dados_variavel(self, nova_lista):
        conecta = conectar_banco_nuvem()
        try:
            from datetime import datetime, timedelta
            data_atual = datetime.now()
            data_atual = data_banco_para_brasileiro(data_atual)

            cursor = conecta.cursor()

            # ================================
            # BUSCA TODOS OS ATIVOS
            # ================================
            cursor.execute("""
                           SELECT id, ticker, nome_pregao, saldo, status, classe_acao
                           FROM variavel_cadastro_ativo
                           WHERE status = 'A';
                           """)

            ativos = cursor.fetchall()

            # ================================
            # BUSCA TODOS MOVIMENTOS
            # ================================
            cursor.execute("""
                           SELECT comp_vend.id_ativo,
                                  comp_vend.qtde,
                                  comp_vend.valor,
                                  mov.data
                           FROM variavel_compra_venda as comp_vend
                                    INNER JOIN movimentacao as mov
                                               ON comp_vend.id_movimentacao = mov.id
                           ORDER BY comp_vend.id_ativo, mov.data;
                           """)

            movimentos_raw = cursor.fetchall()

            # ================================
            # ORGANIZA POR ATIVO
            # ================================
            movimentos_por_ativo = {}

            for id_ativo, qtde, valor, data in movimentos_raw:
                if id_ativo not in movimentos_por_ativo:
                    movimentos_por_ativo[id_ativo] = []

                movimentos_por_ativo[id_ativo].append((qtde, valor, data))

            # ================================
            # PROCESSA ATIVOS
            # ================================
            for ativo in ativos:

                id_ativo, ticker, nome, saldo_db, status, classe = ativo

                movimentos = movimentos_por_ativo.get(id_ativo, [])

                preco_medio, saldo_calculado, data_inicio = self.calcular_preco_medio(movimentos)

                data_inicio = data_banco_para_brasileiro(data_inicio)

                total_medio = saldo_calculado * preco_medio
                total_medio_arred = float_para_moeda_reais(round(total_medio, 2))

                rs_atual = self.extrair_preco_acao(ticker, classe)

                if classe == "CRIPTOMOEDA":
                    dolar = self.obter_dolar_brl()
                    rs_atual_br = rs_atual * dolar

                    totalrs_atual = saldo_calculado * rs_atual_br
                else:
                    totalrs_atual = saldo_calculado * rs_atual

                if rs_atual is None:
                    print(f"⚠️ preço não encontrado para {ticker}")
                    rs_atual = Decimal("0")

                total_rs_atual_arred = float_para_moeda_reais(round(totalrs_atual, 2))

                if classe == "CRIPTOMOEDA":
                    dados = (data_inicio, "BITSO", "VARIÁVEL", ticker, total_medio_arred, data_atual, total_rs_atual_arred, "DIÁRIA", status)
                    nova_lista.append(dados)
                else:
                    dados = (data_inicio, "CLEAR", "VARIÁVEL", ticker, total_medio_arred, data_atual,
                             total_rs_atual_arred, "DIÁRIA", status)
                    nova_lista.append(dados)

                if saldo_db != saldo_calculado:
                    print("⚠️ ERRO: saldo divergente!")

            if nova_lista:
                lanca_tabela(self.table_Lista, nova_lista)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conecta' in locals():
                conecta.close()

    def calcular_totais(self):
        try:
            dados_tabela = extrair_tabela(self.table_Lista)

            total_total = 0
            total_diaria = 0
            total_vencimento = 0

            total_variavel = 0
            total_fixo = 0
            total_fgts = 0

            if dados_tabela:
                for i in dados_tabela:
                    data_mov, banco, tipo, nome, valor, venc, saldo, liq, status = i

                    saldo_float = valores_para_float(saldo)

                    total_total += saldo_float

                    if liq == "DIÁRIA":
                        total_diaria += saldo_float
                    else:
                        total_vencimento += saldo_float

                    if tipo == "VARIÁVEL":
                        total_variavel += saldo_float
                    if tipo == "FGTS":
                        total_fgts += saldo_float
                    if tipo == "FIXO":
                        total_fixo += saldo_float

                    total_diaria_arred = float_para_moeda_reais(round(total_diaria, 2))
                    self.label_Diaria.setText(total_diaria_arred)

                    total_vencimento_arred = float_para_moeda_reais(round(total_vencimento, 2))
                    self.label_Vencimento.setText(total_vencimento_arred)

                    total_total_arred = float_para_moeda_reais(round(total_total, 2))
                    self.label_Total.setText(total_total_arred)

                    total_variavel_arred = float_para_moeda_reais(round(total_variavel, 2))
                    self.label_Variavel.setText(total_variavel_arred)

                    total_fgts_arred = float_para_moeda_reais(round(total_fgts, 2))
                    self.label_FGTS.setText(total_fgts_arred)

                    total_fixo_arred = float_para_moeda_reais(round(total_fixo, 2))
                    self.label_Fixo.setText(total_fixo_arred)


        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def calcular_preco_medio(self, movimentos):
        try:
            saldo = 0
            custo_total = 0
            preco_medio = 0
            data_inicio_posicao = None

            for qtde, valor, data in movimentos:
                saldo_anterior = saldo
                saldo += qtde

                # abriu nova posição
                if saldo_anterior == 0 and saldo > 0:
                    custo_total = 0
                    data_inicio_posicao = data

                # zerou posição
                if saldo == 0:
                    custo_total = 0
                    preco_medio = 0
                    data_inicio_posicao = None
                    continue

                custo_total += valor
                preco_medio = custo_total / saldo

            return preco_medio, saldo, data_inicio_posicao

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def extrair_preco_acao(self, ticker, classe):

        conecta = conectar_banco_nuvem()

        try:

            ticker_codigo = ticker.strip().upper()

            # ===== MONTA TICKER =====
            if classe == "CRIPTOMOEDA":
                if "-" not in ticker_codigo:
                    ticker_codigo += "-USD"

            else:  # AÇÕES e FII
                if not ticker_codigo.endswith(".SA"):
                    ticker_codigo += ".SA"

            t = yf.Ticker(ticker_codigo)

            dados = t.history(period="1d")
            print(dados)

            if dados.empty:
                print(f"Sem dados para {ticker_codigo}")
                return None

            preco_atual = dados["Close"].iloc[-1]
            preco_atual_arred = Decimal(str(round(float(preco_atual), 2)))

            return preco_atual_arred

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conecta' in locals():
                conecta.close()

    def obter_dolar_brl(self):
        try:
            t = yf.Ticker("USDBRL=X")
            dados = t.history(period="1d")

            if dados.empty:
                return Decimal("0")

            preco = dados["Close"].iloc[-1]
            return Decimal(str(round(float(preco), 4)))

        except Exception as e:
            import inspect, sys
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gerar_excel(self):
        try:
            import os
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            def moeda_para_float(valor):
                if valor is None or valor == "":
                    return 0.0
                if isinstance(valor, str):
                    valor = valor.replace("R$", "").strip()
                    valor = valor.replace(".", "").replace(",", ".")
                return float(valor)

            dados_tabela = extrair_tabela(self.table_Lista)

            if not dados_tabela:
                return

            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            arquivo_excel = os.path.join(desktop, "Relatório Ativos.xlsx")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Movimentos"

            # ===== estilos =====
            bold_font = Font(bold=True, color="FFFFFF")
            fill_cinza = PatternFill("solid", fgColor="808080")
            alinhamento_centro = Alignment(horizontal="center", vertical="center")

            borda_preta = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000")
            )

            # ===== cabeçalho =====
            headers = [
                "Data", "Banco", "Tipo", "Nome",
                "Valor", "Vencimento", "Saldo", "Líquido", "Status"
            ]

            ws.append(headers)

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = bold_font
                cell.fill = fill_cinza
                cell.alignment = alinhamento_centro
                cell.border = borda_preta

            # ===== dados =====
            for linha_idx, i in enumerate(dados_tabela, start=2):
                data_mov, banco, tipo, nome, valor, venc, saldo, liq, status = i

                valor_float = moeda_para_float(valor)
                saldo_float = moeda_para_float(saldo)

                ws.append([
                    data_mov,
                    banco,
                    tipo,
                    nome,
                    valor_float,
                    venc,
                    saldo_float,
                    liq,
                    status
                ])

                ws.cell(row=linha_idx, column=5).number_format = '[$R$-416] #,##0.00'
                ws.cell(row=linha_idx, column=7).number_format = '[$R$-416] #,##0.00'

            # ===== bordas + centralizar tudo =====
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                    min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = borda_preta
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # ===== auto largura =====
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2

            wb.save(arquivo_excel)
            self.label_Excel.setText("Excel Salvo!")

        except Exception as e:
            import inspect, sys
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaRelatorioAtivos()
    tela.show()
    qt.exec_()