import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from banco_dados.conexao_nuvem import conectar_banco_nuvem


def calcular_proximas_trocas():
    conecta = conectar_banco_nuvem()
    try:
        cursor = conecta.cursor()

        cursor.execute("""
            SELECT mv.id_veiculo,
                   p.id AS id_produto,
                   p.descricao,
                   mv.km_atual AS km_troca,
                   m.data AS data_troca,
                   p.prazo_km,
                   p.prazo_meses
            FROM manutencao_veiculo_produto mvp
            INNER JOIN manutencao_veiculo mv ON mv.id = mvp.id_manut
            INNER JOIN movimentacao m ON m.id = mv.id_movimentacao
            INNER JOIN cadastro_produto_veiculo p ON p.id = mvp.id_produto
            WHERE p.prazo_km IS NOT NULL OR p.prazo_meses IS NOT NULL
            ORDER BY mv.id_veiculo, p.id, m.data DESC
        """)
        trocas = cursor.fetchall()

        cursor.execute("""
            SELECT id_veiculo, MAX(km_atual) as km_atual
            FROM (
                SELECT id_veiculo, km_atual FROM controle_abastecimento
                UNION ALL
                SELECT id_veiculo, km_atual FROM manutencao_veiculo
            ) t
            GROUP BY id_veiculo
        """)
        km_atuais = {v: km for v, km in cursor.fetchall()}

        resultado = []
        ja_processados = set()

        for v, prod, desc, km_troca, data_troca, prazo_km, prazo_meses in trocas:
            chave = (v, prod)
            if chave in ja_processados:
                continue
            ja_processados.add(chave)

            km_atual = km_atuais.get(v, km_troca)

            proxima_km = km_troca + prazo_km if prazo_km else None
            falta_km = proxima_km - km_atual if proxima_km else None
            proxima_data = (data_troca + timedelta(days=prazo_meses * 30)) if prazo_meses else None
            falta_dias = (proxima_data - datetime.now().date()).days if proxima_data else None

            resultado.append({
                "veiculo": v,
                "produto": desc,
                "ultima_km": km_troca,
                "km_cadastrada": prazo_km,
                "ultima_data": data_troca,
                "proxima_km": proxima_km,
                "falta_km": falta_km,
                "proxima_data": proxima_data,
                "falta_dias": falta_dias
            })

        return resultado

    finally:
        if 'conecta' in locals():
            conecta.close()


def salvar_excel(dados):
    wb = Workbook()
    ws = wb.active
    ws.title = "Trocas"

    # Estilos
    negrito = Font(bold=True)
    borda = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alinhamento = Alignment(horizontal='center', vertical='center')

    # Cabeçalho
    headers = ["Veículo", "Produto", "Últ. KM", "KM Cadastrada", "Últ. Data",
               "Próx. KM", "Falta KM", "Validade até", "Falta Dias"]
    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = negrito
        cell.border = borda
        cell.alignment = alinhamento

    # Linhas
    for r in dados:
        ultima_data_fmt = r['ultima_data'].strftime("%d/%m/%Y") if r['ultima_data'] else "-"
        proxima_data_fmt = r['proxima_data'].strftime("%d/%m/%Y") if r['proxima_data'] else "-"

        linha = [
            r['veiculo'],
            r['produto'],
            r['ultima_km'],
            r['km_cadastrada'] if r['km_cadastrada'] else "-",
            ultima_data_fmt,
            r['proxima_km'] if r['proxima_km'] else "-",
            r['falta_km'] if r['falta_km'] else "-",
            proxima_data_fmt,
            r['falta_dias'] if r['falta_dias'] else "-"
        ]
        ws.append(linha)

    # Aplicar bordas e alinhamento em todas as células
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = borda
            cell.alignment = alinhamento

    # Ajuste automático das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    # Caminho na área de trabalho
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    caminho_arquivo = os.path.join(desktop, f"relatorio_trocas.xlsx")

    wb.save(caminho_arquivo)
    print(f"Relatório salvo em: {caminho_arquivo}")


# ---------------- USO ----------------
if __name__ == "__main__":
    relatorio = calcular_proximas_trocas()
    salvar_excel(relatorio)
