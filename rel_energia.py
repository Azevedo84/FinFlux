import sys
from forms.tela_rel_energia import *
from banco_dados.conexao_nuvem import conectar_banco_nuvem
from banco_dados.controle_erros import grava_erro_banco
from comandos.telas import tamanho_aplicacao, icone
from comandos.tabelas import layout_cabec_tab, lanca_tabela, extrair_tabela
from comandos.conversores import valores_para_float, valores_para_virgula, float_para_moeda_reais, valores_para_int
from PyQt5.QtWidgets import QMainWindow, QApplication, QMessageBox
import inspect
import os
from datetime import datetime
import traceback


class TelaRelatorioEnergia(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        self.processando = False

        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "gremio.png")
        tamanho_aplicacao(self)

        layout_cabec_tab(self.table_Lista)

        self.configurar_data_atual()

        self.btn_Consulta.clicked.connect(self.manipula_dados)
        self.btn_Excel.clicked.connect(self.gerar_excel)

        self.line_Ano.editingFinished.connect(self.manipula_dados)

    def mensagem_alerta(self, mensagem):
        try:
            alert = QMessageBox()
            alert.setIcon(QMessageBox.Warning)
            alert.setText(mensagem)
            alert.setWindowTitle("Atenção")
            alert.setStandardButtons(QMessageBox.Ok)
            alert.exec_()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            self.mensagem_alerta(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                                 f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                                 f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def configurar_data_atual(self):
        try:
            # Obter o mês e o ano atuais
            data_atual = datetime.now()
            mes_atual = data_atual.month  # Exemplo: 11 para novembro
            ano_atual = data_atual.year  # Exemplo: 2024

            # Configurar o ano atual no QLineEdit
            self.line_Ano.setText(str(ano_atual))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tudo(self):
        try:
            self.table_Lista.setRowCount(0)
            self.label_total_custo_km.clear()
            self.label_total_km_lt.clear()
            self.label_total_litros.clear()
            self.label_total_km_rodados.clear()
            self.label_Valor_Total.clear()
            self.label_medio_por_lt.clear()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados(self):
        conecta = conectar_banco_nuvem()

        if not self.processando:
            try:
                self.processando = True

                meses = self.combo_Meses.currentText()
                ano = self.line_Ano.text()

                classifica = self.combo_Classifica.currentText()

                if classifica == "MAIOR VALOR":
                    order_by = "ORDER BY total DESC"
                elif classifica == "MENOR VALOR":
                    order_by = "ORDER BY total"
                elif classifica == "MAIOR CONSUMO":
                    order_by = "ORDER BY energ.valor_consumo DESC"
                elif classifica == "MENOR CONSUMO":
                    order_by = "ORDER BY energ.valor_consumo"
                elif classifica == "MAIOR KWH":
                    order_by = "ORDER BY kwh DESC"
                elif classifica == "MENOR KWH":
                    order_by = "ORDER BY kwh"
                else:
                    order_by = "ORDER BY mov.data"

                self.limpa_tudo()

                if meses and ano:
                    meses_tete = meses.find(" - ")
                    num_mes = meses[:meses_tete]

                    ano_int = int(ano)

                    print("uee")

                    self.total_mensal(num_mes, ano_int, order_by)

                elif ano and not meses:
                    ano_int = int(ano)

                    self.total_anual(ano_int, order_by)

            except Exception as e:
                nome_funcao = inspect.currentframe().f_code.co_name
                exc_traceback = sys.exc_info()[2]
                self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

            finally:
                self.processando = False
                if 'conexao' in locals():
                    conecta.close()

    def select_padrao(self):
        try:
            texto_padrao = """SELECT DATE_FORMAT(mov.data, '%d/%m/%Y'), user.descricao, banc.descricao, 
            DATE_FORMAT(energ.data_anterior, '%d/%m/%Y'), energ.leitura_anterior, 
            DATE_FORMAT(energ.data_atual, '%d/%m/%Y'), energ.leitura_atual, 
            (energ.leitura_atual - energ.leitura_anterior) as kwh, energ.valor_consumo, 
            energ.bandeira, IFNULL(energ.valor_bandeira, ''), 
            IFNULL(energ.outras_taxas, ''), 
            (energ.valor_consumo + IFNULL(energ.valor_bandeira, 0) + IFNULL(energ.outras_taxas, 0)) as total, 
            IFNULL(energ.obs, '')
            FROM controle_energia as energ
            INNER JOIN movimentacao AS mov ON energ.id_movimentacao = mov.id 
            INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id 
            INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id
            INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id 
            INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id """

            return texto_padrao

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def calculo_padrao(self, lista_completa):
        try:
            nova_lista_tabela = []

            litros_final = 0
            total_final = 0
            km_rodados_final = 0

            if lista_completa:
                km_anterior = 0
                for i in lista_completa:
                    data, veiculo, combust, km, litros, unit, total, estab, obs = i

                    total_float = valores_para_float(total)

                    if km_anterior:
                        km_rodados = int(valores_para_int(km) - km_anterior)
                    else:
                        km_rodados = ""

                    if km_rodados:
                        km_lt = round((km_rodados / valores_para_float(litros)), 1)
                        custo_km = float_para_moeda_reais(total_float / km_rodados)
                    else:
                        km_lt = ""
                        custo_km = ""

                    km_anterior = valores_para_float(km)

                    unit_float = valores_para_float(unit)

                    litros_virg = valores_para_virgula(litros)

                    unit_moeda = float_para_moeda_reais(unit_float)
                    total_moeda = float_para_moeda_reais(total)

                    litros_final += valores_para_float(litros)
                    total_final += total_float
                    km_rodados_final += valores_para_float(km_rodados)

                    dados = (data, veiculo, combust, km, litros_virg, unit_moeda, total_moeda, estab, obs, km_rodados,
                             km_lt, custo_km)
                    nova_lista_tabela.append(dados)

            if nova_lista_tabela:
                lanca_tabela(self.table_Lista, nova_lista_tabela)

            if total_final:
                total_arred = float_para_moeda_reais(round(total_final, 2))
                self.label_Valor_Total.setText(str(total_arred))

                litros_arred = valores_para_virgula(round(litros_final, 2))
                self.label_total_litros.setText(litros_arred)

                km_rodado_arred = int(round(km_rodados_final, 2))
                self.label_total_km_rodados.setText(str(km_rodado_arred))

                medio_litro = float_para_moeda_reais(round(total_final, 2) / round(litros_final, 2))
                self.label_medio_por_lt.setText(str(medio_litro))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def total_mensal(self, num_mes, ano_int, order_by):
        conecta = conectar_banco_nuvem()
        try:
            texto_padrao = self.select_padrao()

            cursor = conecta.cursor()
            cursor.execute(f"""{texto_padrao}
                        WHERE MONTH(mov.data) = {num_mes}
                          AND YEAR(mov.data) = {ano_int} 
                          {order_by};
                    """)
            lista_completa = cursor.fetchall()

            if lista_completa:
                lanca_tabela(self.table_Lista, lista_completa)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def total_anual(self, ano_int, order_by):
        conecta = conectar_banco_nuvem()
        try:
            texto_padrao = self.select_padrao()

            cursor = conecta.cursor()
            cursor.execute(f"""{texto_padrao}
                        WHERE YEAR(mov.data) = {ano_int} 
                        {order_by};
                    """)
            lista_completa = cursor.fetchall()

            if lista_completa:
                lanca_tabela(self.table_Lista, lista_completa)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def gerar_excel(self):
        conecta = conectar_banco_nuvem()
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            dados_tabela = extrair_tabela(self.table_Lista)

            if dados_tabela:
                meses = self.combo_Meses.currentText()
                ano = self.line_Ano.text()

                meses_tete = meses.find(" - ")
                num_mes = meses[:meses_tete]

                # Caminho para salvar na área de trabalho
                desktop = os.path.join(os.path.expanduser("~"), "Desktop")
                arquivo_excel = os.path.join(desktop, f"despesas {num_mes} de {ano}.xlsx")

                # Cria um novo workbook e planilha
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Armazem 9"

                # Estilos
                bold_font = Font(bold=True, color="FFFFFF")
                fill_cinza = PatternFill("solid", fgColor="808080")
                alinhamento_centro = Alignment(horizontal="center", vertical="center")
                borda_preta = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000")
                )

                # Cabeçalho
                headers = ["Data", "Banco", "Categoria", "Valor", "Estabelecimento", "Cidade", "Observação"]
                ws.append(headers)

                # Aplicar estilo no cabeçalho
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = bold_font
                    cell.fill = fill_cinza
                    cell.alignment = alinhamento_centro
                    cell.border = borda_preta

                # Adiciona os dados
                for linha_idx, (data, banco, categoria, valor, estab, cidade, obs) in enumerate(dados_tabela, start=2):
                    valor_float = valores_para_float(valor)

                    ws.append([data, banco, categoria, valor_float, estab, cidade, obs])

                    # Formata coluna de valor como moeda (R$)
                    valor_cell = ws.cell(row=linha_idx, column=4)
                    valor_cell.number_format = '[$R$-416] #,##0.00'

                # Aplica borda e alinhamento a todos os dados
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                    for cell in row:
                        cell.border = borda_preta
                        cell.alignment = Alignment(vertical="center")

                # Ajustar largura das colunas automaticamente
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_length + 2

                # Salva o arquivo na área de trabalho
                wb.save(arquivo_excel)
                self.label_Excel.setText(f"Excel Salvo!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conecta' in locals():
                conecta.close()

    def gerar_excel(self):
        conecta = conectar_banco_nuvem()
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            dados_tabela = extrair_tabela(self.table_Lista)

            if dados_tabela:
                meses = self.combo_Meses.currentText()
                ano = self.line_Ano.text()

                meses_tete = meses.find(" - ")
                num_mes = meses[:meses_tete]

                # Caminho para salvar na área de trabalho
                desktop = os.path.join(os.path.expanduser("~"), "Desktop")
                arquivo_excel = os.path.join(desktop, f"despesas {num_mes} de {ano}.xlsx")

                # Cria um novo workbook e planilha
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Armazem 9"

                # Estilos
                bold_font = Font(bold=True, color="FFFFFF")
                fill_cinza = PatternFill("solid", fgColor="808080")
                alinhamento_centro = Alignment(horizontal="center", vertical="center")
                borda_preta = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000")
                )

                # Cabeçalho
                headers = ["Data", "Banco", "Categoria", "Valor", "Estabelecimento", "Cidade", "Observação"]
                ws.append(headers)

                # Aplicar estilo no cabeçalho
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = bold_font
                    cell.fill = fill_cinza
                    cell.alignment = alinhamento_centro
                    cell.border = borda_preta

                # Adiciona os dados
                for linha_idx, (data, banco, categoria, valor, estab, cidade, obs) in enumerate(dados_tabela, start=2):
                    valor_float = valores_para_float(valor)

                    ws.append([data, banco, categoria, valor_float, estab, cidade, obs])

                    # Formata coluna de valor como moeda (R$)
                    valor_cell = ws.cell(row=linha_idx, column=4)
                    valor_cell.number_format = '[$R$-416] #,##0.00'

                # Aplica borda e alinhamento a todos os dados
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                    for cell in row:
                        cell.border = borda_preta
                        cell.alignment = Alignment(vertical="center")

                # Ajustar largura das colunas automaticamente
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_length + 2

                # Salva o arquivo na área de trabalho
                wb.save(arquivo_excel)
                self.label_Excel.setText(f"Excel Salvo!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conecta' in locals():
                conecta.close()


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaRelatorioEnergia()
    tela.show()
    qt.exec_()
