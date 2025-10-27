import sys
from forms.tela_rel_mensal_anual import *
from banco_dados.conexao_nuvem import conectar_banco_nuvem
from banco_dados.controle_erros import grava_erro_banco
from comandos.telas import tamanho_aplicacao, icone
from comandos.tabelas import layout_cabec_tab, lanca_tabela, extrair_tabela
from comandos.conversores import valores_para_float
from PyQt5.QtWidgets import QMainWindow, QApplication, QMessageBox
from PyQt5.QtChart import QChart, QChartView, QPieSeries, QPieSlice
from PyQt5.QtGui import QPainter, QColor, QFont
from PyQt5.QtCore import Qt
import inspect
import os
from datetime import datetime
import traceback


class TelaRelatorioMovMensal(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)
        
        self.processando = False

        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "gremio.png")
        tamanho_aplicacao(self)

        layout_cabec_tab(self.table_Lista)

        self.id_usuario = "1"

        self.lanca_combo_grupo()
        self.configurar_data_atual()

        self.combo_Grupo.activated.connect(self.lanca_combo_categoria)
        self.btn_Consulta.clicked.connect(self.manipula_dados)
        self.btn_Excel.clicked.connect(self.gerar_excel)
        
        self.line_Ano.editingFinished.connect(self.manipula_dados)
        
    def mensagem_alerta(self, mensagem):
        try:
            alert = QMessageBox()
            alert.setIcon(QMessageBox.Warning)
            alert.setText(mensagem)
            alert.setWindowTitle("Atenção")
            alert.setStandardButtons(QMessageBox.Ok)
            alert.exec_()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            self.mensagem_alerta(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                                 f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                                 f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def configurar_data_atual(self):
        try:
            # Obter o mês e o ano atuais
            data_atual = datetime.now()
            mes_atual = data_atual.month  # Exemplo: 11 para novembro
            ano_atual = data_atual.year  # Exemplo: 2024

            # Encontrar o índice do mês atual no combo box e selecioná-lo
            for i in range(self.combo_Meses.count()):
                item_text = self.combo_Meses.itemText(i)
                if item_text.startswith(f"{mes_atual} -"):
                    self.combo_Meses.setCurrentIndex(i)
                    break

            # Configurar o ano atual no QLineEdit
            self.line_Ano.setText(str(ano_atual))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_combo_grupo(self):
        conecta = conectar_banco_nuvem()
        try:
            self.combo_Grupo.clear()

            nova_lista = [""]

            cursor = conecta.cursor()
            cursor.execute("SELECT id, descricao FROM cadastro_grupo "
                           "WHERE id NOT IN (1, 2, 14) "
                           "order by descricao;")
            lista_completa = cursor.fetchall()
            for ides, descr in lista_completa:
                dd = f"{ides} - {descr}"
                nova_lista.append(dd)

            self.combo_Grupo.addItems(nova_lista)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def lanca_combo_categoria(self):
        conecta = conectar_banco_nuvem()
        try:
            self.combo_Categoria.clear()

            nova_lista = [""]

            grupo = self.combo_Grupo.currentText()
            if grupo:
                grupotete = grupo.find(" - ")
                id_grupo = grupo[:grupotete]

                cursor = conecta.cursor()
                cursor.execute(f'SELECT id, descricao FROM cadastro_categoria '
                               f'where id_grupo = {id_grupo} '
                               f'order by descricao;')
                lista_completa = cursor.fetchall()
                for ides, descr in lista_completa:
                    dd = f"{ides} - {descr}"
                    nova_lista.append(dd)

                self.combo_Categoria.addItems(nova_lista)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def manipula_dados(self):
        conecta = conectar_banco_nuvem()

        if not self.processando:
            try:
                self.processando = True

                meses = self.combo_Meses.currentText()
                ano = self.line_Ano.text()

                grupo = self.combo_Grupo.currentText()

                categoria = self.combo_Categoria.currentText()

                classifica = self.combo_Classifica.currentText()

                if classifica == "MAIOR VALOR":
                    order_by = "ORDER BY mov.qtde_sai DESC"
                elif classifica == "MENOR VALOR":
                    order_by = "ORDER BY mov.qtde_sai"
                elif classifica == "GRUPO":
                    order_by = "ORDER BY gr.descricao"
                elif classifica == "CATEGORIA":
                    order_by = "ORDER BY cat.descricao"
                elif classifica == "ESTABELECIMENTO":
                    order_by = "ORDER BY estab.descricao"
                elif classifica == "CIDADE":
                    order_by = "ORDER BY cit.descricao"
                else:
                    order_by = "ORDER BY mov.data"

                if meses and ano:
                    if not grupo:
                        meses_tete = meses.find(" - ")
                        num_mes = meses[:meses_tete]

                        ano_int = int(ano)

                        self.total_mensal(num_mes, ano_int, order_by)
                        self.adicionar_grafico(num_mes, ano_int)
                    else:
                        meses_tete = meses.find(" - ")
                        num_mes = meses[:meses_tete]

                        ano_int = int(ano)

                        grupo_tete = grupo.find(" - ")
                        num_grupo = grupo[:grupo_tete]

                        categoria_tete = categoria.find(" - ")
                        num_categoria = categoria[:categoria_tete]

                        if categoria:
                            self.total_mensal_categoria(num_mes, ano_int, num_categoria, order_by)
                            self.adicionar_grafico(num_mes, ano_int)
                        else:
                            self.total_mensal_grupo(num_mes, ano_int, num_grupo, order_by)
                            self.adicionar_grafico(num_mes, ano_int)

                elif ano and not meses:
                    if not grupo:
                        ano_int = int(ano)

                        self.total_anual(ano_int, order_by)
                        self.adicionar_grafico_ano(ano_int)
                    else:
                        ano_int = int(ano)

                        grupo_tete = grupo.find(" - ")
                        num_grupo = grupo[:grupo_tete]

                        categoria_tete = categoria.find(" - ")
                        num_categoria = categoria[:categoria_tete]

                        if categoria:
                            self.total_anual_categoria(ano_int, num_categoria, order_by)
                            self.adicionar_grafico_ano(ano_int)
                        else:
                            self.total_anual_grupo(ano_int, num_grupo, order_by)
                            self.adicionar_grafico_ano(ano_int)

            except Exception as e:
                nome_funcao = inspect.currentframe().f_code.co_name
                exc_traceback = sys.exc_info()[2]
                self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

            finally:
                self.processando = False
                if 'conexao' in locals():
                    conecta.close()

    def select_padrao(self):
        try:
            texto_padrao = """SELECT mov.id, DATE_FORMAT(mov.data, '%d/%m/%Y') AS data_formatada,
                               banc.descricao, cat.descricao,
                               mov.qtde_sai,
                               estab.descricao, cit.descricao, IFNULL(mov.obs, '')
                        FROM movimentacao AS mov
                        INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id
                        INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id
                        INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id
                        INNER JOIN cadastro_tipoconta AS tip ON sald.id_tipoconta = tip.id
                        INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id
                        INNER JOIN cadastro_grupo AS gr ON cat.id_grupo = gr.id
                        INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id
                        INNER JOIN cadastro_cidade AS cit ON mov.id_cidade = cit.id """

            return texto_padrao

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def total_mensal(self, num_mes, ano_int, order_by):
        conecta = conectar_banco_nuvem()
        try:
            self.table_Lista.setRowCount(0)
            
            texto_padrao = self.select_padrao()

            cursor = conecta.cursor()
            cursor.execute(f"""{texto_padrao}
                        WHERE MONTH(mov.data) = {num_mes}
                          AND YEAR(mov.data) = {ano_int} 
                          AND gr.id NOT IN (1, 2, 14) AND cat.id NOT IN (103, 158, 110) 
                        {order_by};
                    """)
            lista_completa = cursor.fetchall()

            total = 0

            if lista_completa:
                for i in lista_completa:
                    valor = i[4]

                    saldo_float = valores_para_float(valor)

                    total += saldo_float

                lanca_tabela(self.table_Lista, lista_completa)

            if total:
                total_arred = round(total, 2)
                self.label_Total_Lista.setText(str(total_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def total_mensal_grupo(self, num_mes, ano_int, num_grupo, order_by):
        conecta = conectar_banco_nuvem()
        try:
            self.table_Lista.setRowCount(0)

            texto_padrao = self.select_padrao()

            cursor = conecta.cursor()
            cursor.execute(f"""{texto_padrao}
                        WHERE MONTH(mov.data) = {num_mes}
                          AND YEAR(mov.data) = {ano_int} 
                          AND gr.id NOT IN (1, 2, 14) AND cat.id NOT IN (103, 158, 110) 
                          AND gr.id = {num_grupo} 
                        {order_by};
                    """)
            lista_completa = cursor.fetchall()

            total = 0

            if lista_completa:
                for i in lista_completa:
                    valor = i[4]

                    saldo_float = valores_para_float(valor)

                    total += saldo_float

                lanca_tabela(self.table_Lista, lista_completa)

            if total:
                total_arred = round(total, 2)
                self.label_Total_Lista.setText(str(total_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def total_mensal_categoria(self, num_mes, ano_int, num_categoria, order_by):
        conecta = conectar_banco_nuvem()
        try:
            self.table_Lista.setRowCount(0)

            texto_padrao = self.select_padrao()

            cursor = conecta.cursor()
            cursor.execute(f"""{texto_padrao}
                        WHERE MONTH(mov.data) = {num_mes}
                          AND YEAR(mov.data) = {ano_int} 
                          AND cat.id = {num_categoria} 
                        {order_by};
                    """)
            lista_completa = cursor.fetchall()

            total = 0

            if lista_completa:
                for i in lista_completa:
                    valor = i[4]

                    saldo_float = valores_para_float(valor)

                    total += saldo_float

                lanca_tabela(self.table_Lista, lista_completa)

            if total:
                total_arred = round(total, 2)
                self.label_Total_Lista.setText(str(total_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def adicionar_grafico(self, num_mes, ano_int):
        conecta = conectar_banco_nuvem()
        try:
            # Criar a série de dados
            series = QPieSeries()

            cursor = conecta.cursor()
            cursor.execute(f"""
                SELECT gr.descricao, 
                       SUM(mov.qtde_sai) AS total_valor
                FROM movimentacao AS mov
                INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id
                INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id
                INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id
                INNER JOIN cadastro_tipoconta AS tip ON sald.id_tipoconta = tip.id
                INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id
                INNER JOIN cadastro_grupo AS gr ON cat.id_grupo = gr.id
                INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id
                INNER JOIN cadastro_cidade AS cit ON mov.id_cidade = cit.id
                WHERE MONTH(mov.data) = {num_mes}
                  AND YEAR(mov.data) = {ano_int}
                  AND gr.id NOT IN (1, 2, 14) AND cat.id NOT IN (103, 158, 110) 
                GROUP BY gr.descricao
                ORDER BY total_valor DESC;
            """)
            lista_totais = cursor.fetchall()
            if lista_totais:
                # Definir uma paleta de cores
                cores = [
                    QColor("#FF5733"), QColor("#33FF57"), QColor("#3357FF"),
                    QColor("#FF33A1"), QColor("#A133FF"), QColor("#FF8333"),
                    QColor("#33FFF5"), QColor("#F533FF")
                ]

                total = 0

                for i, (categoria, valor) in enumerate(lista_totais):
                    valor_float = valores_para_float(valor)

                    total += valor_float

                    fatia = series.append(categoria, valor_float)

                    # Atribuir uma cor da paleta, usando o índice de forma cíclica
                    fatia.setBrush(cores[i % len(cores)])

                    # Exibir rótulos fora da fatia com descrição e valor
                    fatia.setLabelPosition(QPieSlice.LabelOutside)
                    fatia.setLabel(f"{categoria}: {valor_float:.2f}")
                    fatia.setLabelVisible(True)
                    fatia.setLabelFont(QFont("Arial", 6))  # Define fonte menor (8pt)

                    if valor_float / total < 0.05:  # menos de 5%
                        fatia.setLabelVisible(False)

                # Criar o gráfico
                chart = QChart()
                chart.addSeries(series)

                # Configurar a legenda
                chart.legend().setVisible(True)  # Tornar a legenda visível
                chart.legend().setAlignment(Qt.AlignRight)  # Posicionar a legenda à direita
                chart.legend().setFont(QFont("Arial", 7))

                # Configurar o ChartView
                chart_view = QChartView(chart)
                chart_view.setRenderHint(QPainter.Antialiasing)

                # Adicionar o ChartView ao widget da interface
                layout = self.widget_grafico.layout()  # Obtém o layout do widget
                if not layout:  # Se o layout não existir, cria um
                    from PyQt5.QtWidgets import QVBoxLayout
                    layout = QVBoxLayout(self.widget_grafico)
                else:  # Remove gráficos antigos do layout
                    for i in reversed(range(layout.count())):
                        layout.itemAt(i).widget().deleteLater()

                layout.addWidget(chart_view)  # Adiciona o gráfico ao layout

                if total:
                    total_arred = round(total, 2)
                    self.label_Total_Grafico.setText(str(total_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def total_anual(self, ano_int, order_by):
        conecta = conectar_banco_nuvem()
        try:
            self.table_Lista.setRowCount(0)

            texto_padrao = self.select_padrao()

            cursor = conecta.cursor()
            cursor.execute(f"""{texto_padrao}
                        WHERE YEAR(mov.data) = {ano_int} 
                          AND gr.id NOT IN (1, 2, 14) AND cat.id NOT IN (103, 158, 110) 
                        {order_by};
                    """)
            lista_completa = cursor.fetchall()

            total = 0

            if lista_completa:
                for i in lista_completa:
                    valor = i[4]

                    saldo_float = valores_para_float(valor)

                    total += saldo_float

                lanca_tabela(self.table_Lista, lista_completa)

            if total:
                total_arred = round(total, 2)
                self.label_Total_Lista.setText(str(total_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def total_anual_grupo(self, ano_int, num_grupo, order_by):
        conecta = conectar_banco_nuvem()
        try:
            self.table_Lista.setRowCount(0)

            texto_padrao = self.select_padrao()

            cursor = conecta.cursor()
            cursor.execute(f"""{texto_padrao}
                        WHERE YEAR(mov.data) = {ano_int} 
                          AND gr.id NOT IN (1, 2, 14)  AND cat.id NOT IN (103, 158, 110) 
                          AND gr.id = {num_grupo} 
                        {order_by};
                    """)
            lista_completa = cursor.fetchall()

            total = 0

            if lista_completa:
                for i in lista_completa:
                    valor = i[4]

                    saldo_float = valores_para_float(valor)

                    total += saldo_float

                lanca_tabela(self.table_Lista, lista_completa)

            if total:
                total_arred = round(total, 2)
                self.label_Total_Lista.setText(str(total_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def total_anual_categoria(self, ano_int, num_categoria, order_by):
        conecta = conectar_banco_nuvem()
        try:
            self.table_Lista.setRowCount(0)

            texto_padrao = self.select_padrao()

            cursor = conecta.cursor()
            cursor.execute(f"""{texto_padrao}
                        WHERE YEAR(mov.data) = {ano_int} 
                          AND cat.id = {num_categoria} 
                        {order_by};
                    """)
            lista_completa = cursor.fetchall()

            total = 0

            if lista_completa:
                for i in lista_completa:
                    valor = i[4]

                    saldo_float = valores_para_float(valor)

                    total += saldo_float

                lanca_tabela(self.table_Lista, lista_completa)

            if total:
                total_arred = round(total, 2)
                self.label_Total_Lista.setText(str(total_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def adicionar_grafico_ano(self, ano_int):
        conecta = conectar_banco_nuvem()
        try:
            # Criar a série de dados
            series = QPieSeries()

            cursor = conecta.cursor()
            cursor.execute(f"""
                SELECT gr.descricao, 
                       SUM(mov.qtde_sai) AS total_valor
                FROM movimentacao AS mov
                INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id
                INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id
                INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id
                INNER JOIN cadastro_tipoconta AS tip ON sald.id_tipoconta = tip.id
                INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id
                INNER JOIN cadastro_grupo AS gr ON cat.id_grupo = gr.id
                INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id
                INNER JOIN cadastro_cidade AS cit ON mov.id_cidade = cit.id
                WHERE YEAR(mov.data) = {ano_int}
                  AND gr.id NOT IN (1, 2, 14) AND cat.id NOT IN (103, 158, 110) 
                GROUP BY gr.descricao
                ORDER BY total_valor DESC;
            """)
            lista_totais = cursor.fetchall()
            if lista_totais:
                # Definir uma paleta de cores
                cores = [
                    QColor("#FF5733"), QColor("#33FF57"), QColor("#3357FF"),
                    QColor("#FF33A1"), QColor("#A133FF"), QColor("#FF8333"),
                    QColor("#33FFF5"), QColor("#F533FF")
                ]

                total = 0

                for i, (categoria, valor) in enumerate(lista_totais):
                    valor_float = valores_para_float(valor)

                    total += valor_float

                    fatia = series.append(categoria, valor_float)

                    # Atribuir uma cor da paleta, usando o índice de forma cíclica
                    fatia.setBrush(cores[i % len(cores)])

                    # Exibir rótulos fora da fatia com descrição e valor
                    fatia.setLabelPosition(QPieSlice.LabelOutside)
                    fatia.setLabel(f"{categoria}: {valor_float:.2f}")
                    fatia.setLabelVisible(True)

                # Criar o gráfico
                chart = QChart()
                chart.addSeries(series)

                # Configurar a legenda
                chart.legend().setVisible(True)  # Tornar a legenda visível
                chart.legend().setAlignment(Qt.AlignRight)  # Posicionar a legenda à direita

                # Configurar o ChartView
                chart_view = QChartView(chart)
                chart_view.setRenderHint(QPainter.Antialiasing)

                # Adicionar o ChartView ao widget da interface
                layout = self.widget_grafico.layout()  # Obtém o layout do widget
                if not layout:  # Se o layout não existir, cria um
                    from PyQt5.QtWidgets import QVBoxLayout
                    layout = QVBoxLayout(self.widget_grafico)
                else:  # Remove gráficos antigos do layout
                    for i in reversed(range(layout.count())):
                        layout.itemAt(i).widget().deleteLater()

                layout.addWidget(chart_view)  # Adiciona o gráfico ao layout

                if total:
                    total_arred = round(total, 2)
                    self.label_Total_Grafico.setText(str(total_arred))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def gerar_excel(self):
        conecta = conectar_banco_nuvem()
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            dados_tabela = extrair_tabela(self.table_Lista)

            if dados_tabela:
                meses = self.combo_Meses.currentText()
                ano = self.line_Ano.text()

                meses_tete = meses.find(" - ")
                num_mes = meses[:meses_tete]

                # Caminho para salvar na área de trabalho
                desktop = os.path.join(os.path.expanduser("~"), "Desktop")
                arquivo_excel = os.path.join(desktop, f"despesas {num_mes} de {ano}.xlsx")

                # Cria um novo workbook e planilha
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Despesas"

                # Estilos
                bold_font = Font(bold=True, color="FFFFFF")
                fill_cinza = PatternFill("solid", fgColor="808080")
                alinhamento_centro = Alignment(horizontal="center", vertical="center")
                borda_preta = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000")
                )

                # Cabeçalho
                headers = ["Data", "Usuário", "Banco", "Grupo", "Categoria", "Valor", "Estabelecimento", "Cidade", "Observação"]
                ws.append(headers)

                # Aplicar estilo no cabeçalho
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = bold_font
                    cell.fill = fill_cinza
                    cell.alignment = alinhamento_centro
                    cell.border = borda_preta

                # Adiciona os dados
                for linha_idx, (id_mov, data, banco, categoria, valor, estab, cidade, obs) in enumerate(dados_tabela, start=2):
                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT mov.id, user.descricao "
                                   f"FROM movimentacao AS mov "
                                   f"INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id "
                                   f"INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id "
                                   f"WHERE mov.id = {id_mov};")
                    lista_completa = cursor.fetchall()

                    nome_user = lista_completa[0][1]

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT grup.id, grup.descricao "
                                   f"FROM cadastro_categoria as cat "
                                   f"INNER JOIN cadastro_grupo as grup ON cat.id_grupo = grup.id "
                                   f"where cat.descricao = '{categoria}';")
                    lista_completa = cursor.fetchall()

                    nome_grupo = lista_completa[0][1]

                    valor_float = valores_para_float(valor)

                    ws.append([data, nome_user, banco, nome_grupo, categoria, valor_float, estab, cidade, obs])

                    # Formata coluna de valor como moeda (R$)
                    valor_cell = ws.cell(row=linha_idx, column=4)
                    valor_cell.number_format = '[$R$-416] #,##0.00'

                # Aplica borda e alinhamento a todos os dados
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                    for cell in row:
                        cell.border = borda_preta
                        cell.alignment = Alignment(vertical="center")

                # Ajustar largura das colunas automaticamente
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_length + 2

                # Salva o arquivo na área de trabalho
                wb.save(arquivo_excel)
                self.label_Excel.setText(f"Excel Salvo!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conecta' in locals():
                conecta.close()


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaRelatorioMovMensal()
    tela.show()
    qt.exec_()
