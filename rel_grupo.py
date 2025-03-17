import sys
from comandos.conversores import valores_para_float, float_para_moeda_reais
from forms.tela_rel_grupo import *
from conexao_nuvem import conectar_banco_nuvem
from funcao_padrao import lanca_tabela
from comandos.notificacao import tratar_notificar_erros
from comandos.tabelas import lanca_tabela, layout_cabec_tab, limpa_tabela, extrair_tabela, mensagem_alerta
from comandos.telas import tamanho_aplicacao, icone, cor_fundo_tela
from PyQt5.QtWidgets import QMainWindow, QApplication
import inspect
import os
from datetime import datetime


from comandos.excel import edita_alinhamento, edita_bordas, linhas_colunas_p_edicao
from comandos.excel import criar_workbook, edita_fonte, edita_preenchimento, letra_coluna
from pathlib import Path
from openpyxl.styles import Font, PatternFill


class TelaMovGrupo(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        cor_fundo_tela(self)
        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "gremio.png")
        tamanho_aplicacao(self)

        self.btn_Consulta.clicked.connect(self.manipula_dados)
        self.line_Ano.editingFinished.connect(self.manipula_dados)

        self.btn_Excel.clicked.connect(self.gerar_excel)

        self.line_Ano.setText("2024")

        self.configurar_data_atual()

        self.ano_int = 0
        self.mes_int = 0

    def configurar_data_atual(self):
        try:
            # Obter o mês e o ano atuais
            data_atual = datetime.now()
            mes_atual = data_atual.month  # Exemplo: 11 para novembro
            ano_atual = data_atual.year  # Exemplo: 2024

            # Encontrar o índice do mês atual no combo box e selecioná-lo
            for i in range(self.combo_Meses.count()):
                item_text = self.combo_Meses.itemText(i)
                if item_text.startswith(f"{mes_atual} -"):
                    self.combo_Meses.setCurrentIndex(i)
                    break

            # Configurar o ano atual no QLineEdit
            self.line_Ano.setText(str(ano_atual))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_dados(self):
        try:
            msg = f''
            self.label_msg_excel.setText(msg)

            meses = self.combo_Meses.currentText()
            ano = self.line_Ano.text()

            if meses and ano:
                meses_tete = meses.find(" - ")
                self.mes_int = meses[:meses_tete]

                self.ano_int = int(ano)

                valor_fixa = self.manipula_dados_contas_fixas(self.table_ContaFixa)
                valor_mercado = self.manipula_dados_supermercado(self.table_Supermercado)
                valor_alimento = self.manipula_dados_alimentacao(self.table_Alimentacao)
                valor_assinatura = self.manipula_dados_assinatura(self.table_Assinatura)
                valor_estudos = self.manipula_dados_estudos(self.table_Estudos)
                valor_saude = self.manipula_dados_saude(self.table_Saude)
                valor_lazer = self.manipula_dados_lazer(self.table_Lazer)
                valor_casa = self.manipula_dados_casa(self.table_Casa)
                valor_pets = self.manipula_dados_pets(self.table_Pets)
                valor_impostos = self.manipula_dados_impostos(self.table_Impostos)
                valor_pessoais = self.manipula_dados_pessoais(self.table_Pessoais)
                valor_veiculo = self.manipula_dados_veiculo(self.table_Veiculo)

                total = (valor_fixa + valor_mercado + valor_alimento + valor_assinatura + valor_estudos +
                         valor_saude + valor_lazer + valor_casa + valor_pets + valor_impostos + valor_pessoais +
                         valor_veiculo)

                if total:
                    total_arred = round(total, 2)
                    total_reais = float_para_moeda_reais(total_arred)
                    self.label_Total.setText(str(total_reais))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_dados_contas_fixas(self, tabela):
        conecta = conectar_banco_nuvem()
        try:
            limpa_tabela(tabela)
            layout_cabec_tab(tabela)

            total = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT DATE_FORMAT(mov.data, '%d/%m/%Y') AS data_formatada, "
                           f"banc.descricao, cat.descricao, "
                           f"CASE WHEN mov.qtde_sai = 0 THEN '' ELSE mov.qtde_sai END AS qtde_sai, "
                           f"estab.descricao, IFNULL(mov.obs, '') "
                           f"FROM movimentacao AS mov "
                           f"INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id "
                           f"INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id "
                           f"INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id "
                           f"INNER JOIN cadastro_tipoconta AS tip ON sald.id_tipoconta = tip.id "
                           f"INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id "
                           f"INNER JOIN cadastro_grupo AS gr ON cat.id_grupo = gr.id "
                           f"INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id "
                           f"INNER JOIN cadastro_cidade AS cit ON mov.id_cidade = cit.id "
                           f"WHERE MONTH(mov.data) = {self.mes_int} "
                           f"AND YEAR(mov.data) = {self.ano_int} "
                           f"and gr.id = 4 "
                           f"and cat.id NOT IN (23) "
                           f"ORDER BY mov.data;")
            lista_completa = cursor.fetchall()

            if lista_completa:
                for i in lista_completa:
                    qtde = i[3]

                    qtde_float = valores_para_float(qtde)

                    total += qtde_float

                lanca_tabela(tabela, lista_completa, altura_linha=20, fonte_texto=7)

            if total:
                total_arred = round(total, 2)
                total_reais = float_para_moeda_reais(total_arred)
                self.label_Fixa.setText(str(total_reais))

            return total

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_dados_supermercado(self, tabela):
        conecta = conectar_banco_nuvem()
        try:
            limpa_tabela(tabela)
            layout_cabec_tab(tabela)

            total = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT DATE_FORMAT(mov.data, '%d/%m/%Y') AS data_formatada, "
                           f"banc.descricao, "
                           f"CASE WHEN mov.qtde_sai = 0 THEN '' ELSE mov.qtde_sai END AS qtde_sai, "
                           f"estab.descricao, IFNULL(mov.obs, '') "
                           f"FROM movimentacao AS mov "
                           f"INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id "
                           f"INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id "
                           f"INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id "
                           f"INNER JOIN cadastro_tipoconta AS tip ON sald.id_tipoconta = tip.id "
                           f"INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id "
                           f"INNER JOIN cadastro_grupo AS gr ON cat.id_grupo = gr.id "
                           f"INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id "
                           f"INNER JOIN cadastro_cidade AS cit ON mov.id_cidade = cit.id "
                           f"WHERE MONTH(mov.data) = {self.mes_int} "
                           f"AND YEAR(mov.data) = {self.ano_int} "
                           f"and cat.id = 23 "
                           f"ORDER BY mov.data;")
            lista_completa = cursor.fetchall()

            if lista_completa:
                for i in lista_completa:
                    qtde = i[2]

                    qtde_float = valores_para_float(qtde)

                    total += qtde_float

                lanca_tabela(tabela, lista_completa, altura_linha=20, fonte_texto=7)

            if total:
                total_arred = round(total, 2)
                total_reais = float_para_moeda_reais(total_arred)
                self.label_Mercado.setText(str(total_reais))

            return total

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_dados_alimentacao(self, tabela):
        conecta = conectar_banco_nuvem()
        try:
            limpa_tabela(tabela)
            layout_cabec_tab(tabela)

            total = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT DATE_FORMAT(mov.data, '%d/%m/%Y') AS data_formatada, "
                           f"banc.descricao, cat.descricao, "
                           f"CASE WHEN mov.qtde_sai = 0 THEN '' ELSE mov.qtde_sai END AS qtde_sai, "
                           f"estab.descricao, IFNULL(mov.obs, '') "
                           f"FROM movimentacao AS mov "
                           f"INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id "
                           f"INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id "
                           f"INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id "
                           f"INNER JOIN cadastro_tipoconta AS tip ON sald.id_tipoconta = tip.id "
                           f"INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id "
                           f"INNER JOIN cadastro_grupo AS gr ON cat.id_grupo = gr.id "
                           f"INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id "
                           f"INNER JOIN cadastro_cidade AS cit ON mov.id_cidade = cit.id "
                           f"WHERE MONTH(mov.data) = {self.mes_int} "
                           f"AND YEAR(mov.data) = {self.ano_int} "
                           f"and gr.id = 5 "
                           f"ORDER BY mov.data;")
            lista_completa = cursor.fetchall()

            if lista_completa:
                for i in lista_completa:
                    qtde = i[3]

                    qtde_float = valores_para_float(qtde)

                    total += qtde_float

                lanca_tabela(tabela, lista_completa, altura_linha=20, fonte_texto=7)

            if total:
                total_arred = round(total, 2)
                total_reais = float_para_moeda_reais(total_arred)
                self.label_Alimentacao.setText(str(total_reais))

            return total

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_dados_assinatura(self, tabela):
        conecta = conectar_banco_nuvem()
        try:
            limpa_tabela(tabela)
            layout_cabec_tab(tabela)

            total = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT DATE_FORMAT(mov.data, '%d/%m/%Y') AS data_formatada, "
                           f"banc.descricao, cat.descricao, "
                           f"CASE WHEN mov.qtde_sai = 0 THEN '' ELSE mov.qtde_sai END AS qtde_sai, "
                           f"estab.descricao, IFNULL(mov.obs, '') "
                           f"FROM movimentacao AS mov "
                           f"INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id "
                           f"INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id "
                           f"INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id "
                           f"INNER JOIN cadastro_tipoconta AS tip ON sald.id_tipoconta = tip.id "
                           f"INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id "
                           f"INNER JOIN cadastro_grupo AS gr ON cat.id_grupo = gr.id "
                           f"INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id "
                           f"INNER JOIN cadastro_cidade AS cit ON mov.id_cidade = cit.id "
                           f"WHERE MONTH(mov.data) = {self.mes_int} "
                           f"AND YEAR(mov.data) = {self.ano_int} "
                           f"and gr.id = 6 "
                           f"ORDER BY mov.data;")
            lista_completa = cursor.fetchall()

            if lista_completa:
                for i in lista_completa:
                    qtde = i[3]
                    qtde_float = valores_para_float(qtde)

                    total += qtde_float

                lanca_tabela(tabela, lista_completa, altura_linha=20, fonte_texto=7)

            if total:
                total_arred = round(total, 2)
                total_reais = float_para_moeda_reais(total_arred)
                self.label_Assinatura.setText(str(total_reais))

            return total

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_dados_estudos(self, tabela):
        conecta = conectar_banco_nuvem()
        try:
            limpa_tabela(tabela)
            layout_cabec_tab(tabela)

            total = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT DATE_FORMAT(mov.data, '%d/%m/%Y') AS data_formatada, "
                           f"banc.descricao, cat.descricao, "
                           f"CASE WHEN mov.qtde_sai = 0 THEN '' ELSE mov.qtde_sai END AS qtde_sai, "
                           f"estab.descricao, IFNULL(mov.obs, '') "
                           f"FROM movimentacao AS mov "
                           f"INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id "
                           f"INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id "
                           f"INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id "
                           f"INNER JOIN cadastro_tipoconta AS tip ON sald.id_tipoconta = tip.id "
                           f"INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id "
                           f"INNER JOIN cadastro_grupo AS gr ON cat.id_grupo = gr.id "
                           f"INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id "
                           f"INNER JOIN cadastro_cidade AS cit ON mov.id_cidade = cit.id "
                           f"WHERE MONTH(mov.data) = {self.mes_int} "
                           f"AND YEAR(mov.data) = {self.ano_int} "
                           f"and gr.id = 10 "
                           f"ORDER BY mov.data;")
            lista_completa = cursor.fetchall()

            if lista_completa:
                for i in lista_completa:
                    qtde = i[3]
                    qtde_float = valores_para_float(qtde)

                    total += qtde_float

                lanca_tabela(tabela, lista_completa, altura_linha=20, fonte_texto=7)

            if total:
                total_arred = round(total, 2)
                total_reais = float_para_moeda_reais(total_arred)
                self.label_Estudos.setText(str(total_reais))

            return total

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_dados_casa(self, tabela):
        conecta = conectar_banco_nuvem()
        try:
            limpa_tabela(tabela)
            layout_cabec_tab(tabela)

            total = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT DATE_FORMAT(mov.data, '%d/%m/%Y') AS data_formatada, "
                           f"banc.descricao, gr.descricao, cat.descricao, "
                           f"CASE WHEN mov.qtde_sai = 0 THEN '' ELSE mov.qtde_sai END AS qtde_sai, "
                           f"estab.descricao, IFNULL(mov.obs, '') "
                           f"FROM movimentacao AS mov "
                           f"INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id "
                           f"INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id "
                           f"INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id "
                           f"INNER JOIN cadastro_tipoconta AS tip ON sald.id_tipoconta = tip.id "
                           f"INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id "
                           f"INNER JOIN cadastro_grupo AS gr ON cat.id_grupo = gr.id "
                           f"INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id "
                           f"INNER JOIN cadastro_cidade AS cit ON mov.id_cidade = cit.id "
                           f"WHERE MONTH(mov.data) = {self.mes_int} "
                           f"AND YEAR(mov.data) = {self.ano_int} "
                           f"and gr.id = 3 "
                           f"ORDER BY mov.data;")
            lista_completa = cursor.fetchall()

            if lista_completa:
                for i in lista_completa:
                    qtde = i[4]
                    qtde_float = valores_para_float(qtde)

                    total += qtde_float

                lanca_tabela(tabela, lista_completa, altura_linha=20, fonte_texto=7)

            if total:
                total_arred = round(total, 2)
                total_reais = float_para_moeda_reais(total_arred)
                self.label_Casa.setText(str(total_reais))

            return total

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_dados_pets(self, tabela):
        conecta = conectar_banco_nuvem()
        try:
            limpa_tabela(tabela)
            layout_cabec_tab(tabela)

            total = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT DATE_FORMAT(mov.data, '%d/%m/%Y') AS data_formatada, "
                           f"banc.descricao, gr.descricao, cat.descricao, "
                           f"CASE WHEN mov.qtde_sai = 0 THEN '' ELSE mov.qtde_sai END AS qtde_sai, "
                           f"estab.descricao, IFNULL(mov.obs, '') "
                           f"FROM movimentacao AS mov "
                           f"INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id "
                           f"INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id "
                           f"INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id "
                           f"INNER JOIN cadastro_tipoconta AS tip ON sald.id_tipoconta = tip.id "
                           f"INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id "
                           f"INNER JOIN cadastro_grupo AS gr ON cat.id_grupo = gr.id "
                           f"INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id "
                           f"INNER JOIN cadastro_cidade AS cit ON mov.id_cidade = cit.id "
                           f"WHERE MONTH(mov.data) = {self.mes_int} "
                           f"AND YEAR(mov.data) = {self.ano_int} "
                           f"and gr.id = 9 "
                           f"ORDER BY mov.data;")
            lista_completa = cursor.fetchall()

            if lista_completa:
                for i in lista_completa:
                    qtde = i[4]
                    qtde_float = valores_para_float(qtde)

                    total += qtde_float

                lanca_tabela(tabela, lista_completa, altura_linha=20, fonte_texto=7)

            if total:
                total_arred = round(total, 2)
                total_reais = float_para_moeda_reais(total_arred)
                self.label_Pets.setText(str(total_reais))

            return total

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_dados_impostos(self, tabela):
        conecta = conectar_banco_nuvem()
        try:
            limpa_tabela(tabela)
            layout_cabec_tab(tabela)

            total = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT DATE_FORMAT(mov.data, '%d/%m/%Y') AS data_formatada, "
                           f"banc.descricao, cat.descricao, "
                           f"CASE WHEN mov.qtde_sai = 0 THEN '' ELSE mov.qtde_sai END AS qtde_sai, "
                           f"estab.descricao, IFNULL(mov.obs, '') "
                           f"FROM movimentacao AS mov "
                           f"INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id "
                           f"INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id "
                           f"INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id "
                           f"INNER JOIN cadastro_tipoconta AS tip ON sald.id_tipoconta = tip.id "
                           f"INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id "
                           f"INNER JOIN cadastro_grupo AS gr ON cat.id_grupo = gr.id "
                           f"INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id "
                           f"INNER JOIN cadastro_cidade AS cit ON mov.id_cidade = cit.id "
                           f"WHERE MONTH(mov.data) = {self.mes_int} "
                           f"AND YEAR(mov.data) = {self.ano_int} "
                           f"and gr.id = 13 "
                           f"ORDER BY mov.data;")
            lista_completa = cursor.fetchall()

            if lista_completa:
                for i in lista_completa:
                    qtde = i[3]
                    qtde_float = valores_para_float(qtde)

                    total += qtde_float

                lanca_tabela(tabela, lista_completa, altura_linha=20, fonte_texto=7)

            if total:
                total_arred = round(total, 2)
                total_reais = float_para_moeda_reais(total_arred)
                self.label_Impostos.setText(str(total_reais))

            return total

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_dados_saude(self, tabela):
        conecta = conectar_banco_nuvem()
        try:
            limpa_tabela(tabela)
            layout_cabec_tab(tabela)

            total = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT DATE_FORMAT(mov.data, '%d/%m/%Y') AS data_formatada, "
                           f"banc.descricao, cat.descricao, "
                           f"CASE WHEN mov.qtde_sai = 0 THEN '' ELSE mov.qtde_sai END AS qtde_sai, "
                           f"estab.descricao, IFNULL(mov.obs, '') "
                           f"FROM movimentacao AS mov "
                           f"INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id "
                           f"INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id "
                           f"INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id "
                           f"INNER JOIN cadastro_tipoconta AS tip ON sald.id_tipoconta = tip.id "
                           f"INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id "
                           f"INNER JOIN cadastro_grupo AS gr ON cat.id_grupo = gr.id "
                           f"INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id "
                           f"INNER JOIN cadastro_cidade AS cit ON mov.id_cidade = cit.id "
                           f"WHERE MONTH(mov.data) = {self.mes_int} "
                           f"AND YEAR(mov.data) = {self.ano_int} "
                           f"and gr.id = 7 "
                           f"ORDER BY mov.data;")
            lista_completa = cursor.fetchall()

            if lista_completa:
                for i in lista_completa:
                    qtde = i[3]
                    qtde_float = valores_para_float(qtde)

                    total += qtde_float

                lanca_tabela(tabela, lista_completa, altura_linha=20, fonte_texto=7)

            if total:
                total_arred = round(total, 2)
                total_reais = float_para_moeda_reais(total_arred)
                self.label_Saude.setText(str(total_reais))

            return total

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_dados_lazer(self, tabela):
        conecta = conectar_banco_nuvem()
        try:
            limpa_tabela(tabela)
            layout_cabec_tab(tabela)

            total = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT DATE_FORMAT(mov.data, '%d/%m/%Y') AS data_formatada, "
                           f"banc.descricao, cat.descricao, "
                           f"CASE WHEN mov.qtde_sai = 0 THEN '' ELSE mov.qtde_sai END AS qtde_sai, "
                           f"estab.descricao, IFNULL(mov.obs, '') "
                           f"FROM movimentacao AS mov "
                           f"INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id "
                           f"INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id "
                           f"INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id "
                           f"INNER JOIN cadastro_tipoconta AS tip ON sald.id_tipoconta = tip.id "
                           f"INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id "
                           f"INNER JOIN cadastro_grupo AS gr ON cat.id_grupo = gr.id "
                           f"INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id "
                           f"INNER JOIN cadastro_cidade AS cit ON mov.id_cidade = cit.id "
                           f"WHERE MONTH(mov.data) = {self.mes_int} "
                           f"AND YEAR(mov.data) = {self.ano_int} "
                           f"and gr.id = 11 "
                           f"ORDER BY mov.data;")
            lista_completa = cursor.fetchall()

            if lista_completa:
                for i in lista_completa:
                    qtde = i[3]
                    qtde_float = valores_para_float(qtde)

                    total += qtde_float

                lanca_tabela(tabela, lista_completa, altura_linha=20, fonte_texto=7)

            if total:
                total_arred = round(total, 2)
                total_reais = float_para_moeda_reais(total_arred)
                self.label_Lazer.setText(str(total_reais))

            return total

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_dados_pessoais(self, tabela):
        conecta = conectar_banco_nuvem()
        try:
            limpa_tabela(tabela)
            layout_cabec_tab(tabela)

            total = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT DATE_FORMAT(mov.data, '%d/%m/%Y') AS data_formatada, "
                           f"banc.descricao, cat.descricao, "
                           f"CASE WHEN mov.qtde_sai = 0 THEN '' ELSE mov.qtde_sai END AS qtde_sai, "
                           f"estab.descricao, IFNULL(mov.obs, '') "
                           f"FROM movimentacao AS mov "
                           f"INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id "
                           f"INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id "
                           f"INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id "
                           f"INNER JOIN cadastro_tipoconta AS tip ON sald.id_tipoconta = tip.id "
                           f"INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id "
                           f"INNER JOIN cadastro_grupo AS gr ON cat.id_grupo = gr.id "
                           f"INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id "
                           f"INNER JOIN cadastro_cidade AS cit ON mov.id_cidade = cit.id "
                           f"WHERE MONTH(mov.data) = {self.mes_int} "
                           f"AND YEAR(mov.data) = {self.ano_int} "
                           f"and gr.id = 12 "
                           f"ORDER BY mov.data;")
            lista_completa = cursor.fetchall()

            if lista_completa:
                for i in lista_completa:
                    qtde = i[3]
                    qtde_float = valores_para_float(qtde)

                    total += qtde_float

                lanca_tabela(tabela, lista_completa, altura_linha=20, fonte_texto=7)

            if total:
                total_arred = round(total, 2)
                total_reais = float_para_moeda_reais(total_arred)
                self.label_Pessoais.setText(str(total_reais))

            return total

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_dados_veiculo(self, tabela):
        conecta = conectar_banco_nuvem()
        try:
            limpa_tabela(tabela)
            layout_cabec_tab(tabela)

            total = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT DATE_FORMAT(mov.data, '%d/%m/%Y') AS data_formatada, "
                           f"banc.descricao, cat.descricao, "
                           f"CASE WHEN mov.qtde_sai = 0 THEN '' ELSE mov.qtde_sai END AS qtde_sai, "
                           f"estab.descricao, IFNULL(mov.obs, '') "
                           f"FROM movimentacao AS mov "
                           f"INNER JOIN saldo_banco AS sald ON mov.id_saldo = sald.id "
                           f"INNER JOIN cadastro_usuario AS user ON sald.id_usuario = user.id "
                           f"INNER JOIN cadastro_banco AS banc ON sald.id_banco = banc.id "
                           f"INNER JOIN cadastro_tipoconta AS tip ON sald.id_tipoconta = tip.id "
                           f"INNER JOIN cadastro_categoria AS cat ON mov.id_categoria = cat.id "
                           f"INNER JOIN cadastro_grupo AS gr ON cat.id_grupo = gr.id "
                           f"INNER JOIN cadastro_estabelecimento AS estab ON mov.id_estab = estab.id "
                           f"INNER JOIN cadastro_cidade AS cit ON mov.id_cidade = cit.id "
                           f"WHERE MONTH(mov.data) = {self.mes_int} "
                           f"AND YEAR(mov.data) = {self.ano_int} "
                           f"and gr.id = 8 "
                           f"ORDER BY mov.data;")
            lista_completa = cursor.fetchall()

            if lista_completa:
                for i in lista_completa:
                    qtde = i[3]
                    qtde_float = valores_para_float(qtde)

                    total += qtde_float

                lanca_tabela(tabela, lista_completa, altura_linha=20, fonte_texto=7)

            if total:
                total_arred = round(total, 2)
                total_reais = float_para_moeda_reais(total_arred)
                self.label_Veiculo.setText(str(total_reais))

            return total

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def gerar_excel(self):
        try:
            dados_totais = []
            negritos = []
            qtde_mov = 1

            meses = self.combo_Meses.currentText()
            ano = self.line_Ano.text()

            if meses and ano:
                meses_tete = meses.find(" - ")
                mes_texto = meses[meses_tete + 3:]
                mes_texto = mes_texto.capitalize()

                ano_int = int(ano)
            else:
                mes_texto = ""
                ano_int = 0

            tabela_mercado = extrair_tabela(self.table_Supermercado)
            if tabela_mercado:
                qtde_mov += len(tabela_mercado) + 1
                negritos.append(qtde_mov)

                valor_total = 0
                for i in tabela_mercado:
                    data, banco, rs, estab, obs = i

                    rs_float = valores_para_float(rs)
                    valor_total += rs_float

                    dados = (data, "MERCADO", rs, estab, obs)
                    dados_totais.append(dados)

                dadus = ("", "MERCADO", valor_total, "", "")
                dados_totais.append(dadus)

            tabela_alimentacao = extrair_tabela(self.table_Alimentacao)
            if tabela_alimentacao:
                qtde_mov += len(tabela_alimentacao) + 1
                negritos.append(qtde_mov)

                valor_total = 0

                for i in tabela_alimentacao:
                    data, banco, categoria, rs, estab, obs = i

                    rs_float = valores_para_float(rs)
                    valor_total += rs_float

                    dados = (data, categoria, rs, estab, obs)
                    dados_totais.append(dados)

                dadus = ("", "ALIMENTAÇÃO", valor_total, "", "")
                dados_totais.append(dadus)

            tabela_veiculo = extrair_tabela(self.table_Veiculo)
            if tabela_veiculo:
                qtde_mov += len(tabela_veiculo) + 1
                negritos.append(qtde_mov)

                valor_total = 0

                for i in tabela_veiculo:
                    data, banco, categoria, rs, estab, obs = i

                    rs_float = valores_para_float(rs)
                    valor_total += rs_float

                    dados = (data, categoria, rs, estab, obs)
                    dados_totais.append(dados)

                dadus = ("", "VEÍCULO", valor_total, "", "")
                dados_totais.append(dadus)

            tabela_lazer = extrair_tabela(self.table_Lazer)
            if tabela_lazer:
                qtde_mov += len(tabela_lazer) + 1
                negritos.append(qtde_mov)

                valor_total = 0

                for i in tabela_lazer:
                    data, banco, categoria, rs, estab, obs = i

                    rs_float = valores_para_float(rs)
                    valor_total += rs_float

                    dados = (data, categoria, rs, estab, obs)
                    dados_totais.append(dados)

                dadus = ("", "LAZER", valor_total, "", "")
                dados_totais.append(dadus)

            tabela_fixas = extrair_tabela(self.table_ContaFixa)
            if tabela_fixas:
                qtde_mov += len(tabela_fixas) + 1
                negritos.append(qtde_mov)

                valor_total = 0

                for i in tabela_fixas:
                    data, banco, categoria, rs, estab, obs = i

                    rs_float = valores_para_float(rs)
                    valor_total += rs_float

                    dados = (data, categoria, rs, estab, obs)
                    dados_totais.append(dados)

                dadus = ("", "CONTAS FIXAS", valor_total, "", "")
                dados_totais.append(dadus)

            tabela_estudos = extrair_tabela(self.table_Estudos)
            if tabela_estudos:
                qtde_mov += len(tabela_estudos) + 1
                negritos.append(qtde_mov)

                valor_total = 0

                for i in tabela_estudos:
                    data, banco, categoria, rs, estab, obs = i

                    rs_float = valores_para_float(rs)
                    valor_total += rs_float

                    dados = (data, categoria, rs, estab, obs)
                    dados_totais.append(dados)

                dadus = ("", "ESTUDOS", valor_total, "", "")
                dados_totais.append(dadus)

            tabela_casa = extrair_tabela(self.table_Casa)
            if tabela_casa:
                qtde_mov += len(tabela_casa) + 1
                negritos.append(qtde_mov)

                valor_total = 0

                for i in tabela_casa:
                    data, banco, grupo, categoria, rs, estab, obs = i

                    rs_float = valores_para_float(rs)
                    valor_total += rs_float

                    dados = (data, categoria, rs, estab, obs)
                    dados_totais.append(dados)

                dadus = ("", "CASA", valor_total, "", "")
                dados_totais.append(dadus)

            tabela_saude = extrair_tabela(self.table_Saude)
            if tabela_saude:
                qtde_mov += len(tabela_saude) + 1
                negritos.append(qtde_mov)

                valor_total = 0

                for i in tabela_saude:
                    data, banco, categoria, rs, estab, obs = i

                    rs_float = valores_para_float(rs)
                    valor_total += rs_float

                    dados = (data, categoria, rs, estab, obs)
                    dados_totais.append(dados)

                dadus = ("", "SAÚDE", valor_total, "", "")
                dados_totais.append(dadus)

            tabela_assinatura = extrair_tabela(self.table_Assinatura)
            if tabela_assinatura:
                qtde_mov += len(tabela_assinatura) + 1
                negritos.append(qtde_mov)

                valor_total = 0

                for i in tabela_assinatura:
                    data, banco, categoria, rs, estab, obs = i

                    rs_float = valores_para_float(rs)
                    valor_total += rs_float

                    dados = (data, categoria, rs, estab, obs)
                    dados_totais.append(dados)

                dadus = ("", "ASSINATURA", valor_total, "", "")
                dados_totais.append(dadus)

            tabela_pets = extrair_tabela(self.table_Pets)
            if tabela_pets:
                qtde_mov += len(tabela_pets) + 1
                negritos.append(qtde_mov)

                valor_total = 0

                for i in tabela_pets:
                    data, banco, categoria, rs, estab, obs = i

                    rs_float = valores_para_float(rs)
                    valor_total += rs_float

                    dados = (data, categoria, rs, estab, obs)
                    dados_totais.append(dados)

                dadus = ("", "PETS", valor_total, "", "")
                dados_totais.append(dadus)

            tabela_pessoais = extrair_tabela(self.table_Pessoais)
            if tabela_pessoais:
                qtde_mov += len(tabela_pessoais) + 1
                negritos.append(qtde_mov)

                valor_total = 0

                for i in tabela_pessoais:
                    data, banco, categoria, rs, estab, obs = i

                    rs_float = valores_para_float(rs)
                    valor_total += rs_float

                    dados = (data, categoria, rs, estab, obs)
                    dados_totais.append(dados)

                dadus = ("", "PESSOAIS", valor_total, "", "")
                dados_totais.append(dadus)

            tabela_impostos = extrair_tabela(self.table_Impostos)
            if tabela_impostos:
                qtde_mov += len(tabela_impostos) + 1
                negritos.append(qtde_mov)

                valor_total = 0

                for i in tabela_impostos:
                    data, banco, categoria, rs, estab, obs = i

                    rs_float = valores_para_float(rs)
                    valor_total += rs_float

                    dados = (data, categoria, rs, estab, obs)
                    dados_totais.append(dados)

                dadus = ("", "IMPOSTOS", valor_total, "", "")
                dados_totais.append(dadus)

            if not dados_totais:
                mensagem_alerta(f'As tabelas não podem estar vazias!')
            else:
                workbook = criar_workbook()
                sheet = workbook.active
                sheet.title = f"Despesas"

                headers = ["Data", "Categoria", "Valor", "Estabelecimento", "Observação"]
                sheet.append(headers)

                header_row = sheet[1]
                for cell in header_row:
                    edita_fonte(cell, tamanho=11, negrito=True)
                    edita_preenchimento(cell)
                    edita_alinhamento(cell)

                for dados_ex in dados_totais:
                    data, categoria, rs, estab, obs = dados_ex

                    rs_float = valores_para_float(rs)

                    sheet.append([data, categoria, rs_float, estab, obs])

                for cell in linhas_colunas_p_edicao(sheet, 1, sheet.max_row, 1, sheet.max_column):
                    edita_bordas(cell)
                    edita_alinhamento(cell)

                for column in sheet.columns:
                    max_length = 0
                    column_letter = letra_coluna(column[0].column)
                    for cell in column:
                        if isinstance(cell.value, (int, float)):
                            cell_value_str = "{:.3f}".format(cell.value)
                        else:
                            cell_value_str = str(cell.value)
                        if len(cell_value_str) > max_length:
                            max_length = len(cell_value_str)

                    adjusted_width = (max_length + 2)
                    sheet.column_dimensions[column_letter].width = adjusted_width

                for cell in linhas_colunas_p_edicao(sheet, 2, sheet.max_row, 3, 3):
                    cell.number_format = 'R$ #,##0.00'

                # Define a cor de fundo cinza (RGB: D9D9D9)
                fill_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

                for linha in negritos:
                    for coluna in ["B", "C"]:
                        cell = sheet[f"{coluna}{linha}"]
                        cell.font = Font(bold=True)  # Deixa o texto em negrito
                        cell.fill = fill_gray  # Aplica o fundo cinza

                desktop = Path.home() / "Desktop"
                desk_str = str(desktop)
                nome_req = f'\\Desp Grupo {mes_texto} {ano_int}.xlsx'
                caminho = (desk_str + nome_req)

                workbook.save(caminho)

                msg = f'Excel Criado!'
                self.label_msg_excel.setText(msg)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaMovGrupo()
    tela.show()
    qt.exec_()
